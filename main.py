# -*- coding: utf-8 -*-
"""
020 Email download.py
åŠŸèƒ½ï¼š
1) è¯»å– .env ä¸­ QQ é‚®ç®± IMAP å‡­æ®ï¼Œç™»å½•å¹¶æŠ“å–æœ€è¿‘ N å°é‚®ä»¶ã€‚
2) ä»¥é…ç½®çš„å…³é”®è¯ç­›é€‰ä¸¤ç±»é‚®ä»¶ï¼š
   - KEYWORDS["waiting"]      â†’ â€œç­‰å¾…æ‚¨æŸ¥çœ‹â€
   - KEYWORDS["heyu_da"]      â†’ â€œåˆè‚¥å¸‚å’Œè£•è¾¾â€
   å„è‡ªé€‰å–ã€æœ€æ–°ã€‘ä¸€å°ã€‚
3) æå–é€‰ä¸­é‚®ä»¶çš„ HTML æ­£æ–‡ï¼ˆç”¨äºåç»­è§£æè¡¨æ ¼ï¼‰ï¼Œå¹¶ï¼š
   - è‹¥å‘½ä¸­ heyu_da ç±»ï¼Œä¸‹è½½å…¶é™„ä»¶åˆ°ä¿å­˜ç›®å½•ï¼ˆæ–‡ä»¶åè¿½åŠ æ—¶é—´æˆ³ï¼Œä¿ç•™åŸæ‰©å±•åï¼‰ã€‚
4) å°†â€œé€‰ä¸­çš„ subject + æ”¶åˆ°æ—¶é—´ï¼ˆISO8601ï¼‰â€å†™å…¥ä¿å­˜ç›®å½•ä¸‹çš„ mail_meta.jsonï¼Œä¾›åç»­è„šæœ¬è¯»å–ã€‚
5) è§£æ HTML ä¸­é¦–ä¸ªåˆç†è¡¨æ ¼å¹¶å¯¼å‡º Excelï¼ˆç¬¬ä¸€é¡µï¼‰ã€‚
ä½¿ç”¨ï¼š
- å¯ä¼ å…¥ä¿å­˜ç›®å½•ä½œä¸ºç¬¬1ä¸ªå‘½ä»¤è¡Œå‚æ•°ï¼›ä¸ä¼ åˆ™ä½¿ç”¨å¹³å°é»˜è®¤ç›®å½•ï¼ˆWindows: ./dataï¼›å…¶ä»–: ~/dataï¼‰ã€‚
"""

import os
import sys
import re
import time
import platform
import json
import email
import imaplib
from email.header import decode_header
from email.utils import parsedate_tz, mktime_tz
from datetime import datetime

import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from dotenv import load_dotenv

# ================================
# ğŸ“‚ è·¯å¾„é…ç½®ï¼ˆæ”¯æŒä¸»ç¨‹åºä¼ å‚ï¼‰
# ================================
if platform.system() == "Windows":
    default_save_path = os.path.join(os.getcwd(), "data")  # Windows: ç›¸å¯¹è·¯å¾„ ./data
else:
    default_save_path = os.path.expanduser("~/data")       # Linux/macOS: å®¶ç›®å½• ~/data

excel_save_path = sys.argv[1] if len(sys.argv) >= 2 else default_save_path
os.makedirs(excel_save_path, exist_ok=True)
print(f"ğŸ“‚ ä¿å­˜è·¯å¾„: {os.path.abspath(excel_save_path)}")

# ================================
# ğŸ”§ å…³é”®è¯/é‚®ç®±é…ç½®ï¼ˆé›†ä¸­ç®¡ç†ï¼‰
# ================================
KEYWORDS = {
    "waiting": "ç­‰å¾…æ‚¨æŸ¥çœ‹",
    "heyu_da": "åˆè‚¥å¸‚å’Œè£•è¾¾",
}
MAILBOX = os.getenv("IMAP_MAILBOX", "INBOX")          # QQ/å¤šæ•° IMAP å…¼å®¹ "INBOX"
RECENT_LIMIT = int(os.getenv("RECENT_LIMIT", "100"))  # æœ€è¿‘æŠ“å–é‚®ä»¶æ•°é‡ä¸Šé™
META_FILENAME = "mail_meta.json"                      # å…ƒæ•°æ®æ–‡ä»¶åï¼ˆå†™å…¥ excel_save_pathï¼‰

# ================================
# ğŸ“§ é‚®ç®±å‡­æ®ï¼ˆ.envï¼‰
# ================================
load_dotenv()  # å¯æ”¹ä¸º load_dotenv(dotenv_path="...") å®šç‚¹åŠ è½½

email_user = os.getenv("EMAIL_ADDRESS_QQ")
# å…¼å®¹æ—§å†™æ³• EMAIL_PASSWOR_QQï¼ˆå°‘äº†Dï¼‰
email_password = os.getenv("EMAIL_PASSWORD_QQ") or os.getenv("EMAIL_PASSWOR_QQ")
email_server = os.getenv("IMAP_SERVER", "imap.qq.com")

if not email_user or not email_password:
    raise ValueError("âŒ ç¯å¢ƒå˜é‡æœªæ­£ç¡®é…ç½®ï¼ˆEMAIL_ADDRESS_QQ / EMAIL_PASSWORD_QQï¼‰ï¼")

print("ğŸ“¬ æ­£åœ¨ä½¿ç”¨é‚®ç®±:", email_user)

# ================================
# ğŸ”‘ æ ‡é¢˜è§£ç ä¸æ¸…ç†
# ================================
def decode_str(s: str) -> str:
    if not s:
        return ""
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    elif isinstance(value, bytes):
        value = value.decode("utf-8", errors="ignore")
    return value

def clean_subject(subject: str) -> str:
    cleaned_subject = re.sub(r'\[([^\[\]]+)\]', r'\1', subject or "")
    cleaned_subject = re.sub(r'ã€([^ã€ã€‘]+)ã€‘', r'\1', cleaned_subject)
    return cleaned_subject.strip()

# ================================
# ğŸ“¨ æŠ“å–é‚®ä»¶å¹¶è¾“å‡º HTML/å…ƒæ•°æ®/é™„ä»¶
# ================================
def fetch_html_from_emails(server: str, user: str, password: str, save_dir: str) -> str | None:
    mail = None
    html_content = None

    # é¢„ç½®å…ƒæ•°æ®ï¼ˆä¸¤ç±»æœ€æ–°é‚®ä»¶ï¼‰
    meta = {
        "selected_heyu_da_subject": None,
        "selected_heyu_da_received_at": None,
        "selected_waiting_subject": None,
        "selected_waiting_received_at": None,
    }

    try:
        print("ğŸ”— æ­£åœ¨è¿æ¥é‚®ç®±...")
        mail = imaplib.IMAP4_SSL(server)
        mail.login(user, password)

        # é€‰æ‹©é‚®ç®±ç›®å½•
        status, _ = mail.select(MAILBOX)
        if status != "OK":
            print(f"âš ï¸ æ— æ³•é€‰æ‹©é‚®ç®±ç›®å½• {MAILBOX}ï¼Œå°è¯•ä½¿ç”¨ INBOX")
            mail.select("INBOX")

        print(f"ğŸ” æ­£åœ¨æ£€ç´¢æœ€è¿‘ {RECENT_LIMIT} å°é‚®ä»¶...")
        status, messages = mail.search(None, "ALL")
        if status != "OK":
            print("æœªæ‰¾åˆ°é‚®ä»¶")
            return None

        mail_ids = messages[0].split()
        if not mail_ids:
            print("é‚®ç®±ä¸ºç©ºã€‚")
            return None

        recent_mail_ids = mail_ids[-RECENT_LIMIT:]
        print(f"ğŸ“¨ å…± {len(mail_ids)} å°ï¼Œå¤„ç†æœ€è¿‘ {len(recent_mail_ids)} å°ã€‚")

        inventory_query_emails = []

        # éå†æœ€è¿‘çš„ N å°é‚®ä»¶
        for i, mail_id in enumerate(recent_mail_ids, start=1):
            status, msg_data = mail.fetch(mail_id, "(RFC822)")
            if status != "OK" or not msg_data or not msg_data[0]:
                print(f"âš ï¸ ç¬¬ {i} å°æŠ“å–å¤±è´¥")
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = decode_str(msg.get("Subject"))
            from_ = decode_str(msg.get("From"))
            date_raw = decode_str(msg.get("Date"))

            # è½¬æ¢ä¸º datetimeï¼Œå¤±è´¥åˆ™å…œåº•ä¸º 1970-01-01
            mail_date = parsedate_tz(date_raw)
            if mail_date:
                mail_datetime = datetime.fromtimestamp(mktime_tz(mail_date))
            else:
                mail_datetime = datetime(1970, 1, 1)

            cleaned_subject = clean_subject(subject)

            print(f"  Â· ç¬¬ {i} å° | åŸ: {subject} | æ¸…ç†: {cleaned_subject} | å‘ä»¶äºº: {from_} | æ”¶åˆ°: {mail_datetime}")

            # ä»…æ”¶é›†æ ‡é¢˜å‘½ä¸­ä¸¤ç±»å…³é”®è¯çš„é‚®ä»¶
            if (KEYWORDS["waiting"] in cleaned_subject) or (KEYWORDS["heyu_da"] in cleaned_subject):
                inventory_query_emails.append({
                    "mail_id": mail_id,
                    "subject": subject,
                    "cleaned_subject": cleaned_subject,
                    "date": mail_datetime,
                    "msg": msg
                })

        # æ‰“å°ç­›é€‰åˆ—è¡¨
        if inventory_query_emails:
            print("\nâœ… å‘½ä¸­å…³é”®è¯çš„é‚®ä»¶ï¼š")
            for item in inventory_query_emails:
                print(f"  - {item['cleaned_subject']} | {item['date']}")
        else:
            print("\nâ„¹ï¸ æœªå‘½ä¸­ä»»ä½•å…³é”®è¯é‚®ä»¶ã€‚")

        # é€‰å‡ºâ€œåˆè‚¥å¸‚å’Œè£•è¾¾â€æœ€æ–°ä¸€å° â†’ æå– HTML + ä¸‹è½½é™„ä»¶
        selected_heyu = _pick_latest(inventory_query_emails, KEYWORDS["heyu_da"])
        if selected_heyu:
            html_content = extract_html_from_msg(selected_heyu["msg"]) or html_content
            print(f"\nğŸ“Œ é€‰ä¸­(åˆè‚¥å¸‚å’Œè£•è¾¾): {selected_heyu['cleaned_subject']} | {selected_heyu['date']}")
            meta["selected_heyu_da_subject"] = selected_heyu["cleaned_subject"]
            meta["selected_heyu_da_received_at"] = selected_heyu["date"].isoformat() if selected_heyu["date"] else None
            # ä¸‹è½½é™„ä»¶
            download_attachments(selected_heyu["msg"], save_dir)

        # é€‰å‡ºâ€œç­‰å¾…æ‚¨æŸ¥çœ‹â€æœ€æ–°ä¸€å° â†’ æå– HTML
        selected_waiting = _pick_latest(inventory_query_emails, KEYWORDS["waiting"])
        if selected_waiting:
            html_content = extract_html_from_msg(selected_waiting["msg"]) or html_content
            print(f"\nğŸ“Œ é€‰ä¸­(ç­‰å¾…æ‚¨æŸ¥çœ‹): {selected_waiting['cleaned_subject']} | {selected_waiting['date']}")
            meta["selected_waiting_subject"] = selected_waiting["cleaned_subject"]
            meta["selected_waiting_received_at"] = selected_waiting["date"].isoformat() if selected_waiting["date"] else None

        # å†™å‡ºå…ƒæ•°æ®
        _write_meta(meta, os.path.join(save_dir, META_FILENAME))

        if html_content:
            print("âœ… å·²è·å–é€‰å®šé‚®ä»¶çš„ HTML æ­£æ–‡ã€‚")
        else:
            print("â„¹ï¸ æœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„ HTML æ­£æ–‡ã€‚")

        return html_content

    except imaplib.IMAP4.error as e:
        print(f"IMAP é”™è¯¯: {e}")
        return None
    except Exception as e:
        print(f"è·å–é‚®ä»¶å¤±è´¥: {e}")
        return None
    finally:
        try:
            if mail is not None:
                mail.logout()
        except Exception:
            pass

def _pick_latest(candidates: list[dict], keyword: str) -> dict | None:
    """åœ¨ candidates ä¸­é€‰å‡º cleaned_subject å« keyword çš„ã€æœ€æ–°ã€‘ä¸€å°ã€‚"""
    selected = None
    for item in candidates:
        if keyword in item["cleaned_subject"]:
            if (selected is None) or (item["date"] > selected["date"]):
                selected = item
    return selected

# ================================
# ğŸ§© ä»é‚®ä»¶ä¸­æå– HTML æ­£æ–‡
# ================================
def extract_html_from_msg(msg) -> str | None:
    html_content = None
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition") or "")
            if content_type == "text/html" and "attachment" not in content_disposition:
                charset = part.get_content_charset() or part.get_charset() or "utf-8"
                try:
                    html_content = part.get_payload(decode=True).decode(charset, errors="ignore")
                except Exception:
                    html_content = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                break
    else:
        if msg.get_content_type() == "text/html":
            charset = msg.get_content_charset() or msg.get_charset() or "utf-8"
            try:
                html_content = msg.get_payload(decode=True).decode(charset, errors="ignore")
            except Exception:
                html_content = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
    return html_content

# ================================
# ğŸ“ ä¸‹è½½é™„ä»¶ï¼ˆè¿½åŠ æ—¶é—´æˆ³ï¼Œä¿ç•™æ‰©å±•åï¼‰
# ================================
def download_attachments(msg, download_folder: str) -> None:
    if not msg.is_multipart():
        return
    for part in msg.walk():
        content_disposition = str(part.get("Content-Disposition") or "")
        if "attachment" in content_disposition:
            # è§£ç é™„ä»¶æ–‡ä»¶å
            raw_name = part.get_filename()
            if not raw_name:
                continue
            filename, encoding = decode_header(raw_name)[0]
            if isinstance(filename, bytes):
                filename = filename.decode(encoding or "utf-8", errors="ignore")

            base_name, ext = os.path.splitext(filename)
            current_time = time.strftime("%Y%m%d_%H%M%S")
            safe_name = f"{base_name}_{current_time}{ext or ''}"
            file_path = os.path.join(download_folder, safe_name)

            file_data = part.get_payload(decode=True)
            with open(file_path, "wb") as f:
                f.write(file_data)

            print(f"ğŸ“¥ é™„ä»¶å·²ä¸‹è½½: {file_path}")

# ================================
# ğŸ§  è§£æ HTML è¡¨æ ¼å¹¶å¯¼å‡º Excel
# ================================
def parse_html_table(html_content: str) -> list[list[str]]:
    if not html_content:
        return []
    print("ğŸ”§ æ­£åœ¨è§£æ HTML è¡¨æ ¼...")
    soup = BeautifulSoup(html_content, "html.parser")
    table = soup.find("table")
    if not table:
        print("æœªæ‰¾åˆ°ä»»ä½• <table>ã€‚")
        return []

    data = []
    rows = table.find_all("tr")
    header = None

    for idx, row in enumerate(rows):
        cols = [ele.get_text(strip=True) for ele in row.find_all(["td", "th"])]
        if not cols:
            continue

        if header is None:
            # é¦–ä¸ªéç©ºè¡Œï¼Œä½œä¸ºè¡¨å¤´ï¼ˆè‹¥åˆ—æ•°è¿‡äºå¼‚å¸¸å¯ç»§ç»­å¯»æ‰¾ä¸‹ä¸€è¡Œï¼Œè¿™é‡Œæ²¿ç”¨ç®€åŒ–ç­–ç•¥ï¼‰
            header = cols
            data.append(header)
            continue

        # ä¸è¡¨å¤´åˆ—æ•°ä¸ä¸€è‡´ â†’ è·³è¿‡
        if len(cols) != len(header):
            continue
        # é‡å¤è¡¨å¤´ â†’ è·³è¿‡
        if cols == header:
            continue

        data.append(cols)

    print(f"âœ… å·²æå– {len(data)} è¡Œï¼ˆå«è¡¨å¤´ï¼‰ã€‚")
    return data

def save_to_excel(data: list[list[str]], save_dir: str, file_prefix="å­˜é‡æŸ¥è¯¢") -> None:
    if not data:
        print("â„¹ï¸ æ²¡æœ‰å¯å¯¼å‡ºçš„æ•°æ®ã€‚")
        return

    # å»é‡ï¼ˆä¿åºï¼‰
    seen = set()
    unique_data = []
    for row in data:
        tup = tuple(row)
        if tup not in seen:
            seen.add(tup)
            unique_data.append(row)

    df = pd.DataFrame(unique_data)

    # æ–‡ä»¶ååŠ æ—¶é—´æˆ³
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    file_name = f"{file_prefix}_{timestamp}.xlsx"
    full_path = os.path.join(save_dir, file_name)

    print(f"ğŸ’¾ æ­£åœ¨ä¿å­˜ Excelï¼š{full_path}")
    # ä¸å†™ headerï¼ˆåŸé€»è¾‘ï¼‰
    df.to_excel(full_path, index=False, header=False)

    # openpyxl å†ç»†åŒ–æ ¼å¼
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    ws.title = "ç¬¬ä¸€é¡µ"

    # ç¤ºä¾‹ï¼šå°è¯•å¯¹ç¬¬5åˆ—ã€ç¬¬6åˆ—åº”ç”¨åƒåˆ†ä½ & å³å¯¹é½ï¼ˆä»…å½“å¤šæ•°å¯è¢«è¯†åˆ«ä¸ºæ•°å€¼ï¼‰
    decimal_columns = [4, 5]  # 0-based ç´¢å¼•ï¼Œå¯¹åº”ç¬¬5/6åˆ—
    max_row = ws.max_row
    for col in decimal_columns:
        # ç»Ÿè®¡æ•°å€¼æ¯”ä¾‹
        numeric_count = 0
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=col + 1).value
            try:
                float(str(val).replace(",", ""))  # å°è¯•å¯è½¬æ•°
                numeric_count += 1
            except Exception:
                pass
        # è¶…è¿‡ä¸€åŠå¯è§†ä¸ºæ•°å€¼ â†’ åº”ç”¨æ ¼å¼
        if numeric_count >= (max_row - 1) / 2:
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col + 1)
                try:
                    v = float(str(cell.value).replace(",", ""))
                    cell.value = v
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                except Exception:
                    # æ— æ³•è½¬æ•°å€¼å°±è·³è¿‡
                    pass

    wb.save(full_path)
    print("âœ… Excel ä¿å­˜å®Œæˆã€‚")

def _write_meta(meta: dict, path: str) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        print(f"ğŸ“ å…ƒæ•°æ®å·²å†™å…¥: {path}")
    except Exception as e:
        print(f"âš ï¸ å…ƒæ•°æ®å†™å…¥å¤±è´¥: {e}")

# ================================
# ğŸš€ ä¸»ç¨‹åº
# ================================
if __name__ == '__main__':
    print("ç¨‹åºå¯åŠ¨")
    html_content = fetch_html_from_emails(email_server, email_user, email_password, excel_save_path)

    if html_content:
        preview = html_content[:400].replace("\n", " ")
        print(f"HTML é¢„è§ˆ: {preview} ...")

        table_data = parse_html_table(html_content)
        if table_data:
            save_to_excel(table_data, excel_save_path, file_prefix="å­˜é‡æŸ¥è¯¢")
        else:
            print("è¡¨æ ¼ä¸ºç©ºï¼Œæœªå¯¼å‡º Excelã€‚")
    else:
        print("æœªè·å–åˆ° HTMLï¼Œç¨‹åºç»“æŸã€‚")
