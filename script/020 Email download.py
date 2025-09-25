# -*- coding: utf-8 -*-
"""
020 Email download.py  (åŒ—äº¬æ—¶é—´ç‰ˆ)
- ç»Ÿä¸€æ‰€æœ‰æ—¶é—´ä¸º Asia/Shanghaiï¼ˆUTC+8ï¼‰
- mail_meta.json å†…å†™å…¥å¸¦åç§»çš„ ISO8601
"""
import os
import sys
import re
import platform
import json
import email
import imaplib
from email.header import decode_header
from email.utils import parsedate_tz, mktime_tz
from datetime import datetime
from zoneinfo import ZoneInfo  # Python 3.9+

import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from dotenv import load_dotenv

# ================================
# ğŸ•’ æ—¶åŒºå·¥å…·ï¼ˆç»Ÿä¸€åŒ—äº¬æ—¶é—´ï¼‰
# ================================
TZ_SH = ZoneInfo("Asia/Shanghai")

def now_shanghai() -> datetime:
    return datetime.now(TZ_SH)

def ts_to_shanghai(ts: float) -> datetime:
    return datetime.fromtimestamp(ts, tz=TZ_SH)

# ================================
# ğŸ“‚ è·¯å¾„é…ç½®ï¼ˆæ”¯æŒä¸»ç¨‹åºä¼ å‚ï¼‰
# ================================
if platform.system() == "Windows":
    default_save_path = os.path.join(os.getcwd(), "data")
else:
    default_save_path = os.path.expanduser("~/data")

excel_save_path = sys.argv[1] if len(sys.argv) >= 2 else default_save_path
os.makedirs(excel_save_path, exist_ok=True)
print(f"ğŸ“‚ ä¿å­˜è·¯å¾„: {os.path.abspath(excel_save_path)}")

# ================================
# ğŸ”§ å…³é”®è¯/é‚®ç®±é…ç½®ï¼ˆé›†ä¸­ç®¡ç†ï¼‰
# ================================
KEYWORDS = {
    "waiting": "ç­‰å¾…æ‚¨æŸ¥çœ‹",
    "heyu_da": "åˆè‚¥å¸‚å’Œè£•è¾¾",
}
MAILBOX = os.getenv("IMAP_MAILBOX", "INBOX")
RECENT_LIMIT = int(os.getenv("RECENT_LIMIT", "15"))
META_FILENAME = "mail_meta.json"

# ================================
# ğŸ“§ é‚®ç®±å‡­æ®ï¼ˆ.envï¼‰
# ================================
load_dotenv()
email_user = os.getenv("EMAIL_ADDRESS_QQ")
email_password = os.getenv("EMAIL_PASSWORD_QQ") or os.getenv("EMAIL_PASSWOR_QQ")
email_server = os.getenv("IMAP_SERVER", "imap.qq.com")

if not email_user or not email_password:
    raise ValueError("âŒ ç¯å¢ƒå˜é‡æœªæ­£ç¡®é…ç½®ï¼ˆEMAIL_ADDRESS_QQ / EMAIL_PASSWORD_QQï¼‰ï¼")

print("ğŸ“¬ æ­£åœ¨ä½¿ç”¨é‚®ç®±:", email_user)

# ================================
# ğŸ”‘ æ ‡é¢˜è§£ç ä¸æ¸…ç†
# ================================
def decode_str(s: str) -> str:
    if not s:
        return ""
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    elif isinstance(value, bytes):
        value = value.decode("utf-8", errors="ignore")
    return value

def clean_subject(subject: str) -> str:
    cleaned_subject = re.sub(r'\[([^\[\]]+)\]', r'\1', subject or "")
    cleaned_subject = re.sub(r'ã€([^ã€ã€‘]+)ã€‘', r'\1', cleaned_subject)
    return cleaned_subject.strip()

# ================================
# ğŸ“¨ æŠ“å–é‚®ä»¶å¹¶è¾“å‡º HTML/å…ƒæ•°æ®/é™„ä»¶
# ================================
def fetch_html_from_emails(server: str, user: str, password: str, save_dir: str) -> str | None:
    mail = None
    html_content = None

    meta = {
        "selected_heyu_da_subject": None,
        "selected_heyu_da_received_at": None,  # ISO8601ï¼ˆå¸¦+08:00ï¼‰
        "selected_waiting_subject": None,
        "selected_waiting_received_at": None,  # ISO8601ï¼ˆå¸¦+08:00ï¼‰
    }

    try:
        print("ğŸ”— æ­£åœ¨è¿æ¥é‚®ç®±...")
        mail = imaplib.IMAP4_SSL(server)
        mail.login(user, password)

        status, _ = mail.select(MAILBOX)
        if status != "OK":
            print(f"âš ï¸ æ— æ³•é€‰æ‹©é‚®ç®±ç›®å½• {MAILBOX}ï¼Œå°è¯•ä½¿ç”¨ INBOX")
            mail.select("INBOX")

        print(f"ğŸ” æ­£åœ¨æ£€ç´¢æœ€è¿‘ {RECENT_LIMIT} å°é‚®ä»¶...")
        status, messages = mail.search(None, "ALL")
        if status != "OK":
            print("æœªæ‰¾åˆ°é‚®ä»¶")
            return None

        mail_ids = messages[0].split()
        if not mail_ids:
            print("é‚®ç®±ä¸ºç©ºã€‚")
            return None

        recent_mail_ids = mail_ids[-RECENT_LIMIT:]
        print(f"ğŸ“¨ å…± {len(mail_ids)} å°ï¼Œå¤„ç†æœ€è¿‘ {len(recent_mail_ids)} å°ã€‚")

        inventory_query_emails = []

        for i, mail_id in enumerate(recent_mail_ids, start=1):
            status, msg_data = mail.fetch(mail_id, "(RFC822)")
            if status != "OK" or not msg_data or not msg_data[0]:
                print(f"âš ï¸ ç¬¬ {i} å°æŠ“å–å¤±è´¥")
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = decode_str(msg.get("Subject"))
            from_ = decode_str(msg.get("From"))
            date_raw = decode_str(msg.get("Date"))

            mail_date = parsedate_tz(date_raw)
            if mail_date:
                # mktime_tz è¿”å› UTC ç§’æ•°ï¼›ç›´æ¥è½¬æ¢ä¸ºâ€œåŒ—äº¬æ—¶é—´â€ aware datetime
                mail_datetime = ts_to_shanghai(mktime_tz(mail_date))
            else:
                mail_datetime = datetime(1970, 1, 1, tzinfo=TZ_SH)

            cleaned_subject = clean_subject(subject)
            print(f"  Â· ç¬¬ {i} å° | åŸ: {subject} | æ¸…ç†: {cleaned_subject} | å‘ä»¶äºº: {from_} | æ”¶åˆ°(åŒ—äº¬): {mail_datetime.strftime('%Y-%m-%d %H:%M:%S %z')}")

            if (KEYWORDS["waiting"] in cleaned_subject) or (KEYWORDS["heyu_da"] in cleaned_subject):
                inventory_query_emails.append({
                    "mail_id": mail_id,
                    "subject": subject,
                    "cleaned_subject": cleaned_subject,
                    "date": mail_datetime,  # Aware(Asia/Shanghai)
                    "msg": msg
                })

        if inventory_query_emails:
            print("\nâœ… å‘½ä¸­å…³é”®è¯çš„é‚®ä»¶ï¼š")
            for item in inventory_query_emails:
                print(f"  - {item['cleaned_subject']} | {item['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
        else:
            print("\nâ„¹ï¸ æœªå‘½ä¸­ä»»ä½•å…³é”®è¯é‚®ä»¶ã€‚")

        # é€‰å‡ºâ€œåˆè‚¥å¸‚å’Œè£•è¾¾â€æœ€æ–°ä¸€å°
        selected_heyu = _pick_latest(inventory_query_emails, KEYWORDS["heyu_da"])
        if selected_heyu:
            html_content = extract_html_from_msg(selected_heyu["msg"]) or html_content
            print(f"\nğŸ“Œ é€‰ä¸­(åˆè‚¥å¸‚å’Œè£•è¾¾): {selected_heyu['cleaned_subject']} | {selected_heyu['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
            meta["selected_heyu_da_subject"] = selected_heyu["cleaned_subject"]
            meta["selected_heyu_da_received_at"] = selected_heyu["date"].isoformat()
            download_attachments(selected_heyu["msg"], save_dir)

        # é€‰å‡ºâ€œç­‰å¾…æ‚¨æŸ¥çœ‹â€æœ€æ–°ä¸€å°
        selected_waiting = _pick_latest(inventory_query_emails, KEYWORDS["waiting"])
        if selected_waiting:
            html_content = extract_html_from_msg(selected_waiting["msg"]) or html_content
            print(f"\nğŸ“Œ é€‰ä¸­(ç­‰å¾…æ‚¨æŸ¥çœ‹): {selected_waiting['cleaned_subject']} | {selected_waiting['date'].strftime('%Y-%m-%d %H:%M:%S %z')}")
            meta["selected_waiting_subject"] = selected_waiting["cleaned_subject"]
            meta["selected_waiting_received_at"] = selected_waiting["date"].isoformat()

        _write_meta(meta, os.path.join(save_dir, META_FILENAME))

        if html_content:
            print("âœ… å·²è·å–é€‰å®šé‚®ä»¶çš„ HTML æ­£æ–‡ã€‚")
        else:
            print("â„¹ï¸ æœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„ HTML æ­£æ–‡ã€‚")

        return html_content

    except imaplib.IMAP4.error as e:
        print(f"IMAP é”™è¯¯: {e}")
        return None
    except Exception as e:
        print(f"è·å–é‚®ä»¶å¤±è´¥: {e}")
        return None
    finally:
        try:
            if mail is not None:
                mail.logout()
        except Exception:
            pass

def _pick_latest(candidates: list[dict], keyword: str) -> dict | None:
    selected = None
    for item in candidates:
        if keyword in item["cleaned_subject"]:
            if (selected is None) or (item["date"] > selected["date"]):
                selected = item
    return selected

# ================================
# ğŸ§© ä»é‚®ä»¶ä¸­æå– HTML æ­£æ–‡
# ================================
def extract_html_from_msg(msg) -> str | None:
    html_content = None
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition") or "")
            if content_type == "text/html" and "attachment" not in content_disposition:
                charset = part.get_content_charset() or part.get_charset() or "utf-8"
                try:
                    html_content = part.get_payload(decode=True).decode(charset, errors="ignore")
                except Exception:
                    html_content = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                break
    else:
        if msg.get_content_type() == "text/html":
            charset = msg.get_content_charset() or msg.get_charset() or "utf-8"
            try:
                html_content = msg.get_payload(decode=True).decode(charset, errors="ignore")
            except Exception:
                html_content = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
    return html_content

# ================================
# ğŸ“ ä¸‹è½½é™„ä»¶ï¼ˆæ–‡ä»¶åè¿½åŠ â€œåŒ—äº¬æ—¶é—´â€æ—¶é—´æˆ³ï¼‰
# ================================
def download_attachments(msg, download_folder: str) -> None:
    """ä¸‹è½½é‚®ä»¶é™„ä»¶ï¼šæ–‡ä»¶åæŒ‰ åŸå_YYYYMMDD_HHMMSSï¼ˆåŒ—äº¬ï¼‰+æ‰©å±•åã€‚"""
    if not msg.is_multipart():
        return

    import mimetypes
    import unicodedata

    def _decode_filename(raw: str) -> str:
        parts = decode_header(raw)
        s = ""
        for p, enc in parts:
            if isinstance(p, bytes):
                s += p.decode(enc or "utf-8", errors="ignore")
            else:
                s += p
        s = unicodedata.normalize("NFC", s).replace("ï¼", ".").strip().strip(".")
        return s

    def _sanitize(name: str) -> str:
        invalid = '<>:"/\\|?*'
        name = "".join((c if c not in invalid else "_") for c in name).strip().strip(".")
        return name or "attachment"

    def _guess_ext(content_type: str) -> str:
        overrides = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
            "application/vnd.ms-excel": ".xls",
            "text/csv": ".csv",
            "application/zip": ".zip",
            "application/pdf": ".pdf",
        }
        return overrides.get(content_type) or (mimetypes.guess_extension(content_type) or "")

    def _ensure_unique(path: str) -> str:
        if not os.path.exists(path):
            return path
        base, ext = os.path.splitext(path)
        i = 2
        while True:
            candidate = f"{base}_{i}{ext}"
            if not os.path.exists(candidate):
                return candidate
            i += 1

    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue

        content_disposition = str(part.get("Content-Disposition") or "")
        raw_name = part.get_filename()

        if "attachment" not in content_disposition and not raw_name:
            continue

        if raw_name:
            filename = _decode_filename(raw_name)
        else:
            filename = f"attachment{_guess_ext(part.get_content_type())}"

        base_name, ext = os.path.splitext(filename)
        if not ext:
            ext = _guess_ext(part.get_content_type())

        ts = now_shanghai().strftime("%Y%m%d_%H%M%S")  # åŒ—äº¬æ—¶é—´
        safe_base = _sanitize(base_name)
        safe_name = f"{safe_base}_{ts}{ext}"
        file_path = os.path.join(download_folder, safe_name)
        file_path = _ensure_unique(file_path)

        file_data = part.get_payload(decode=True)
        if not file_data:
            continue
        with open(file_path, "wb") as f:
            f.write(file_data)

        print(f"ğŸ“¥ é™„ä»¶å·²ä¸‹è½½(åŒ—äº¬æ—¶): {file_path}")

# ================================
# ğŸ§  è§£æ HTML è¡¨æ ¼å¹¶å¯¼å‡º Excel
# ================================
def parse_html_table(html_content: str) -> list[list[str]]:
    print("æ­£åœ¨è§£æ HTML å†…å®¹ä¸­çš„è¡¨æ ¼...")

    try:
        snap_path = os.path.join(excel_save_path, "last_mail_html.html")
        with open(snap_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"ğŸ” HTML å¿«ç…§: {snap_path}")
    except Exception:
        pass

    soup = BeautifulSoup(html_content, "html.parser")
    table = soup.find("table")
    if not table:
        print("æœªæ‰¾åˆ° HTML è¡¨æ ¼ï¼")
        return []

    data = []
    rows = table.find_all("tr")
    header = None

    for idx, row in enumerate(rows):
        cols = [ele.get_text(strip=True) for ele in row.find_all(["td", "th"])]
        if not cols:
            print(f"ç¬¬ {idx + 1} è¡Œæ˜¯ç©ºè¡Œï¼Œè·³è¿‡")
            continue
        if header is None:
            if idx == 0 and len(cols) > 10:
                print("ç¬¬ä¸€è¡Œåˆ—æ•°è¿‡å¤šï¼Œè®¤ä¸ºå…¶ä¸ºæ­£æ–‡å†…å®¹ï¼Œè·³è¿‡")
                continue
            header = cols
            data.append(header)
            continue
        if len(cols) != len(header):
            print(f"ç¬¬ {idx + 1} è¡Œåˆ—æ•°ä¸è¡¨å¤´ä¸åŒ¹é…ï¼Œè·³è¿‡")
            continue
        if cols == header:
            print(f"ç¬¬ {idx + 1} è¡Œæ˜¯é‡å¤è¡¨å¤´ï¼Œè·³è¿‡")
            continue
        data.append(cols)

    print(f"æˆåŠŸæå– {len(data)} è¡Œè¡¨æ ¼æ•°æ®ã€‚")

    for i in range(len(data)):
        for j in range(len(data[i])):
            if isinstance(data[i][j], str) and data[i][j].isdigit():
                data[i][j] = str(data[i][j])

    return data

def save_to_excel(data: list[list[str]], save_dir: str, file_prefix="å­˜é‡æŸ¥è¯¢") -> None:
    if not data:
        print("â„¹ï¸ æ²¡æœ‰å¯å¯¼å‡ºçš„æ•°æ®ã€‚")
        return

    seen = set()
    unique_data = []
    for row in data:
        tup = tuple(row)
        if tup not in seen:
            seen.add(tup)
            unique_data.append(row)

    df = pd.DataFrame(unique_data)

    # æ–‡ä»¶åç”¨åŒ—äº¬æ—¶é—´
    timestamp = now_shanghai().strftime("%Y%m%d_%H%M%S")
    file_name = f"{file_prefix}_{timestamp}.xlsx"
    full_path = os.path.join(save_dir, file_name)

    print(f"ğŸ’¾ æ­£åœ¨ä¿å­˜ Excelï¼ˆåŒ—äº¬æ—¶ï¼‰: {full_path}")
    df.to_excel(full_path, index=False, header=False)

    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    ws.title = "ç¬¬ä¸€é¡µ"

    decimal_columns = [4, 5]  # ç¬¬5/6åˆ—ï¼ˆ0-basedï¼‰
    max_row = ws.max_row
    for col in decimal_columns:
        numeric_count = 0
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=col + 1).value
            try:
                float(str(val).replace(",", ""))
                numeric_count += 1
            except Exception:
                pass
        if numeric_count >= (max_row - 1) / 2:
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col + 1)
                try:
                    v = float(str(cell.value).replace(",", ""))
                    cell.value = v
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                except Exception:
                    pass

    wb.save(full_path)
    print("âœ… Excel ä¿å­˜å®Œæˆã€‚")

def _write_meta(meta: dict, path: str) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        print(f"ğŸ“ å…ƒæ•°æ®å·²å†™å…¥: {path}")
    except Exception as e:
        print(f"âš ï¸ å…ƒæ•°æ®å†™å…¥å¤±è´¥: {e}")

# ================================
# ğŸš€ ä¸»ç¨‹åº
# ================================
if __name__ == '__main__':
    print(f"ç¨‹åºå¯åŠ¨ï¼ˆåŒ—äº¬æ—¶ï¼‰: {now_shanghai().strftime('%Y-%m-%d %H:%M:%S %z')}")
    html_content = fetch_html_from_emails(email_server, email_user, email_password, excel_save_path)

    if html_content:
        preview = html_content[:400].replace("\n", " ")
        print(f"HTML é¢„è§ˆ: {preview} ...")

        table_data = parse_html_table(html_content)
        if table_data:
            save_to_excel(table_data, excel_save_path, file_prefix="å­˜é‡æŸ¥è¯¢")
        else:
            print("è¡¨æ ¼ä¸ºç©ºï¼Œæœªå¯¼å‡º Excelã€‚")
    else:
        print("æœªè·å–åˆ° HTMLï¼Œç¨‹åºç»“æŸã€‚")
