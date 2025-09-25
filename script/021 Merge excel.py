# -*- coding: utf-8 -*-
"""
021 Merge excel.py
- ä»¥ã€æœ€æ–°çš„ã€‘æ–‡ä»¶ååŒ…å«â€œåˆè‚¥å¸‚â€çš„ Excel ä½œä¸ºåŸºåº•
- æ’é™¤å…¶å®ƒæ‰€æœ‰â€œåˆè‚¥å¸‚â€æ—§æ–‡ä»¶ï¼Œé¿å…è¦†ç›–
- è¾“å‡ºæ–‡ä»¶ååç¼€ä½¿ç”¨åŒ—äº¬æ—¶é—´ï¼ˆUTC+8ï¼‰
"""
import os
import sys
import shutil
import platform
from datetime import datetime
from openpyxl import load_workbook
from zoneinfo import ZoneInfo   # Python 3.9+ å†…ç½®æ—¶åŒºåº“

# ================================
# âš™ï¸ é…ç½®åŒº
# ================================
KEYWORD_BASE = "åˆè‚¥å¸‚"
EXT = ".xlsx"
SORT_OTHERS_BY_MTIME_ASC = True
PRINT_PREFIX = "âœ…"

# ================================
# ğŸ“‚ è·¯å¾„è·å–
# ================================
if platform.system() == "Windows":
    default_folder_path = os.path.join(os.getcwd(), "data")
else:
    default_folder_path = os.path.join(os.getcwd(), "data")

if len(sys.argv) >= 2:
    folder_path = os.path.join(sys.argv[1])
    print(f"{PRINT_PREFIX} ä½¿ç”¨ä¼ å…¥è·¯å¾„: {folder_path}")
else:
    folder_path = default_folder_path
    print(f"âš ï¸ æœªä¼ å…¥è·¯å¾„ï¼Œä½¿ç”¨é»˜è®¤è·¯å¾„: {folder_path}")

if not os.path.exists(folder_path):
    print(f"âŒ æ–‡ä»¶å¤¹ä¸å­˜åœ¨: {folder_path}")
    sys.exit(1)

# æœ‰æ•ˆæ–‡ä»¶è¿‡æ»¤
def is_valid_xlsx(name: str) -> bool:
    return (
        name.endswith(EXT)
        and not os.path.basename(name).startswith("~$")
        and os.path.isfile(os.path.join(folder_path, name))
    )

all_excels = [f for f in os.listdir(folder_path) if is_valid_xlsx(f)]
if not all_excels:
    print("âš ï¸ ç›®å½•ä¸‹æ²¡æœ‰å¯ç”¨çš„ .xlsx æ–‡ä»¶")
    sys.exit(0)

# ================================
# ğŸ” æ‰¾æœ€æ–°çš„â€œåˆè‚¥å¸‚â€æ–‡ä»¶
# ================================
hefei_candidates = [f for f in all_excels if KEYWORD_BASE in f]
print(f"{PRINT_PREFIX} æ‰¾åˆ° {len(hefei_candidates)} ä¸ªæ–‡ä»¶ååŒ…å« '{KEYWORD_BASE}':")
for i, f in enumerate(hefei_candidates, 1):
    print(f"{i}. {f}")

if not hefei_candidates:
    print(f"âŒ æœªæ‰¾åˆ°åŒ…å«â€œ{KEYWORD_BASE}â€çš„åŸºåº•æ–‡ä»¶")
    sys.exit(1)

hefei_candidates = sorted(
    hefei_candidates,
    key=lambda f: os.path.getmtime(os.path.join(folder_path, f)),
    reverse=True
)
base_file = hefei_candidates[0]
base_path = os.path.join(folder_path, base_file)
print(f"\n{PRINT_PREFIX} åŸºåº•æ–‡ä»¶ï¼ˆæœ€æ–°ï¼‰: {base_file}")

# ================================
# ğŸ§º å…¶å®ƒå¾…åˆå¹¶æ–‡ä»¶
# ================================
other_excel_files = [f for f in all_excels if f not in hefei_candidates]
if SORT_OTHERS_BY_MTIME_ASC:
    other_excel_files = sorted(other_excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
else:
    other_excel_files = sorted(other_excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)), reverse=True)

print(f"{PRINT_PREFIX} å¾…åˆå¹¶æ–‡ä»¶æ•°: {len(other_excel_files)}")
for i, f in enumerate(other_excel_files, 1):
    print(f"{i}. {f}")

# ================================
# ğŸ†• åˆ›å»ºåˆå¹¶æ–‡ä»¶ï¼ˆæ–‡ä»¶åç”¨åŒ—äº¬æ—¶é—´ï¼‰
# ================================
beijing_now = datetime.now(ZoneInfo("Asia/Shanghai"))
timestamp = beijing_now.strftime("%Y%m%d_%H%M%S")
merged_filename = f"æ€»åº“å­˜{timestamp}.xlsx"
merged_filepath = os.path.join(folder_path, merged_filename)

shutil.copy(base_path, merged_filepath)
print(f"\n{PRINT_PREFIX} åˆ›å»ºåˆå¹¶æ–‡ä»¶: {merged_filename}")

# ================================
# ğŸ“‘ åˆå¹¶é€»è¾‘
# ================================
def copy_sheet_values(src_wb, dst_wb, sheet_name: str):
    src = src_wb[sheet_name]
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    dst = dst_wb.create_sheet(sheet_name)
    for row in src.iter_rows(values_only=True):
        dst.append(list(row))

merged_wb = load_workbook(merged_filepath)

for file in other_excel_files:
    file_path = os.path.join(folder_path, file)
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        print(f"âš ï¸ è·³è¿‡æ— æ³•æ‰“å¼€çš„æ–‡ä»¶: {file}ï¼ŒåŸå› ï¼š{e}")
        continue

    for sheet_name in wb.sheetnames:
        copy_sheet_values(wb, merged_wb, sheet_name)
        print(f"{PRINT_PREFIX} å¤åˆ¶å·¥ä½œè¡¨: {sheet_name} â† {file}")
    wb.close()

if "Sheet" in merged_wb.sheetnames and len(merged_wb["Sheet"]["A"]) == 0:
    try:
        del merged_wb["Sheet"]
    except Exception:
        pass

merged_wb.save(merged_filepath)
merged_wb.close()
print(f"{PRINT_PREFIX} åˆå¹¶å®Œæˆï¼Œè¾“å‡º: {merged_filepath}")
