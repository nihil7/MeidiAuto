import os
import sys
import re
import glob
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

# =========================================
# ğŸ”§ CONFIGï½œé›†ä¸­é…ç½®ï¼ˆåªæ”¹è¿™é‡Œï¼‰
# -----------------------------------------
CONFIG = {
    # 0) é‚®ä»¶å…ƒæ•°æ®ï¼ˆå†™å…¥ M3 çš„æ—¶é—´æ¥æºï¼‰
    "meta_filename": "mail_meta.json",                  # data ç›®å½•ä¸‹çš„ json æ–‡ä»¶
    "meta_waiting_key": "selected_waiting_received_at", # mail_meta.json ä¸­çš„é”®

    # 1) è·¯å¾„ä¸ç›®æ ‡åŒ¹é…
    "default_folder": os.path.join(os.getcwd(), "data"),   # æœªä¼ å‚æ—¶çš„é»˜è®¤ç›®å½•
    "inventory_pattern": "*æ€»åº“å­˜*.xlsx",                   # åº“å­˜æ–‡ä»¶ååŒ¹é…æ¨¡å¼
    "target_sheet": "åº“å­˜è¡¨",                               # ç›®æ ‡å·¥ä½œè¡¨ï¼ˆè¢«å†™å…¥/æ ¼å¼åŒ–ï¼‰

    # 2) æœ‰æ•ˆæ•°æ®èŒƒå›´ä¸ç»“æ„
    "first_data_row": 4,     # ä»ç¬¬å‡ è¡Œå¼€å§‹è§†ä¸ºè¡¨ä½“ï¼ˆç”¨äºæ‰¾Båˆ—ç¬¬ä¸€ä¸ªç©ºè¡Œï¼‰
    "align_start_row": 5,    # ä»ç¬¬å‡ è¡Œå¼€å§‹åº”ç”¨ä¼šè®¡æ ¼å¼/å¯¹é½
    "insert_after_J_cols": 10,  # åœ¨ J åˆ—åæ’å…¥å¤šå°‘åˆ—
    "insert_after_B_cols": 1,   # åœ¨ B åˆ—åæ’å…¥å¤šå°‘åˆ—ï¼ˆç”¨äºç¼–å·åˆ—ï¼‰

    # 3) ç‰ˆé¢ä¸è§†å›¾
    "freeze_panes": "A5",    # å†»ç»“çª—æ ¼ä½ç½®
    "row1_height": 18,       # ç¬¬1è¡Œè¡Œé«˜
    "hide_row2": True,       # æ˜¯å¦éšè—ç¬¬2è¡Œ
    "zoom_scale": 95,        # æ‰“å¼€ç¼©æ”¾
    "focus_cell": "A5",      # æ‰“å¼€æ—¶èšç„¦å•å…ƒæ ¼ï¼ˆå¹¶æ¿€æ´» target_sheetï¼‰

    # 4) åˆ—å®½ä¸éšè—
    "column_widths": {       # å®½åº¦å•ä½ä¸ºExcelåˆ—å®½ï¼ˆä¼šåœ¨è®¾ç½®æ—¶ +0.6ï¼‰
        "B": 4.5, "C": 0.1, "D": 35.88, "E": 3, "F": 3.6,
        "G": 8.6, "H": 8, "I": 8, "J": 8.8,
        "K": 5.88, "L": 8.1, "M": 9.8, "N": 5.88,
        "O": 9.5, "P": 9.8, "Q": 10.08, "R": 9.5, "S": 9.5
    },
    "hidden_columns": ["C"],     # æ‰“å¼€æ—¶é»˜è®¤éšè—çš„åˆ—ï¼ˆä¸å½±å“è¯»å†™ï¼‰
    "left_align_cols": ["T"],    # éœ€è¦å·¦å¯¹é½çš„åˆ—ï¼ˆåˆ—å­—æ¯å½¢å¼ï¼‰ï¼Œä¾‹ï¼šTåˆ—å·¦å¯¹é½

    # 5) è¡¨å¤´ä¸åˆå¹¶
    "merge_ranges": [("H3", "J3"), ("U3", "W3")],  # åˆå¹¶åŒºåŸŸ
    "set_value_cells": {"U3": "ä¸åˆæ ¼"},           # åˆå¹¶åéœ€è¦å†™å…¥çš„å•å…ƒæ ¼å†…å®¹
    "headers_K_row4": [                            # ä» K4 èµ·å‘å³å†™å…¥çš„è¡¨å¤´
        "å¤–åº”å­˜", "æœ€å°å‘è´§", "å®¶é‡Œåº“å­˜", "å®¶åº”å­˜", "æ’äº§",
        "æœˆè®¡åˆ’", "æœˆè®¡åˆ’ç¼ºå£", "å¤–ä»“å‡ºåº“æ€»é‡", "å¤–ä»“å…¥åº“æ€»é‡", "å¤‡æ³¨"
    ],
    "header_fill_color": "C187F7",                 # è¡¨å¤´å¡«å……è‰²ï¼ˆ16è¿›åˆ¶RGBï¼‰

    # 6) â€œç¬¬ä¸€é¡µå‰¯æœ¬â€ä¸â€œå®¶é‡Œåº“å­˜â€æ´¾ç”Ÿ
    "src_sheet_for_copy": "ç¬¬ä¸€é¡µ",  # è‹¥å­˜åœ¨åˆ™ç­›é€‰ç”Ÿæˆâ€œç¬¬ä¸€é¡µå‰¯æœ¬â€
    "warehouse_col_name": "ä»“åº“",
    "warehouse_keep_value": "æˆå“åº“",
    "copy_sheet_name": "ç¬¬ä¸€é¡µå‰¯æœ¬",
    "home_sheet_name": "å®¶é‡Œåº“å­˜",
    "home_cols": ["ç¼–å·", "å­˜è´§åç§°", "æ•°é‡"],  # å®¶é‡Œåº“å­˜ä¸‰åˆ—è¡¨å¤´
    "home_name_col": "å­˜è´§åç§°",
    "home_qty_col": "ä¸»æ•°é‡",

    # 7) å›å¡«è§„åˆ™ï¼ˆå®¶é‡Œåº“å­˜ â†’ åº“å­˜è¡¨ï¼‰
    "backfill_target_col_index": 13,  # å›å¡«åˆ°åº“å­˜è¡¨çš„åˆ—ç´¢å¼•ï¼ˆMåˆ—=13ï¼‰
    "regex_4digit_dash": r"\d{4}-",   # 'dddd-' ç”¨å4ä½åŒ¹é…
    "regex_5digit": r"\d{5}",         # 5ä½æ ‡å‡†ç¼–å·åŒ¹é…

    # 8) ä¼šè®¡æ ¼å¼ä¸å¯¹é½ï¼ˆG~Qï¼‰
    "acc_fmt_cols": (7, 17),  # åˆ—èŒƒå›´ï¼ˆG=7 ~ Q=17ï¼‰
    "acc_fmt_row_from": 5,    # ä»ç¬¬5è¡Œå¼€å§‹
    "acc_number_format": '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)',

    # 9) å…¶ä»–
    "center_col_letter": "C",  # éœ€è¦å±…ä¸­çš„ç¼–å·åˆ—
    "center_title_cell": "B1", # å·¦å¯¹é½çš„æ ‡é¢˜å•å…ƒæ ¼
    "number_header_cell": "C4" # ç¼–å·åˆ—è¡¨å¤´
}
# =========================================

def _read_waiting_time(folder, meta_name, key):
    """è¯»å– mail_meta.json ä¸­çš„ selected_waiting_received_atï¼ˆå­—ç¬¦ä¸²ï¼‰"""
    meta_path = os.path.join(folder, meta_name)
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
        val = meta.get(key)
        return str(val) if val else None
    except FileNotFoundError:
        print(f"â„¹ï¸ æœªæ‰¾åˆ°å…ƒæ•°æ®æ–‡ä»¶: {meta_path}")
    except Exception as e:
        print(f"âš ï¸ è¯»å–å…ƒæ•°æ®å¤±è´¥: {e}")
    return None

def main(cfg: dict):
    # ---------- è·¯å¾„ ----------
    folder_path = sys.argv[1] if len(sys.argv) >= 2 else cfg["default_folder"]
    if not os.path.exists(folder_path):
        print(f"âŒ è·¯å¾„ä¸å­˜åœ¨: {folder_path}")
        sys.exit(1)

    files = glob.glob(os.path.join(folder_path, cfg["inventory_pattern"]))
    if not files:
        print("âŒ æœªæ‰¾åˆ°åŒ…å«â€œæ€»åº“å­˜â€çš„æ–‡ä»¶")
        sys.exit(1)
    latest_file = max(files, key=os.path.getmtime)

    # ---------- æ‰“å¼€å·¥ä½œç°¿ ----------
    wb = load_workbook(latest_file)
    if cfg["target_sheet"] not in wb.sheetnames:
        print(f"âŒ ç¼ºå°‘å·¥ä½œè¡¨ï¼š{cfg['target_sheet']}")
        sys.exit(1)
    sh = wb[cfg["target_sheet"]]

    # ---------- æ‰¾Båˆ—ç¬¬ä¸€ä¸ªç©ºè¡Œ ----------
    max_row = sh.max_row
    last_empty_row = max_row + 1
    for r in range(cfg["first_data_row"], max_row + 1):
        if sh[f"B{r}"].value is None:
            last_empty_row = r
            break
    print(f"âš¡ Båˆ—ç¬¬ä¸€ä¸ªç©ºè¡Œ: {last_empty_row}")

    # ---------- è§£é™¤åˆå¹¶ ----------
    for rng in list(sh.merged_cells.ranges):
        sh.unmerge_cells(str(rng))

    # ---------- æ’å…¥åˆ— ----------
    sh.insert_cols(10, cfg["insert_after_J_cols"])  # J åæ’å…¥
    sh.insert_cols(3,  cfg["insert_after_B_cols"])  # B åæ’å…¥ï¼ˆç¼–å·åˆ—ï¼‰

    # ---------- å†™å…¥â€œç­‰å¾…æ‚¨æŸ¥çœ‹â€çš„æ”¶åˆ°æ—¶é—´åˆ° M3ï¼Œå¹¶å·¦å¯¹é½ ----------
    waiting_time = _read_waiting_time(folder_path, cfg["meta_filename"], cfg["meta_waiting_key"])
    if waiting_time:
        sh["M3"].value = waiting_time
        sh["M3"].alignment = Alignment(horizontal="left", vertical="center")
        print(f"ğŸ•’ å·²å†™å…¥ M3ï¼ˆç­‰å¾…æ‚¨æŸ¥çœ‹æ—¶é—´ï¼‰: {waiting_time}")
    else:
        print("ğŸ•’ æ²¡æœ‰å¯å†™å…¥çš„ç­‰å¾…æ—¶é—´ï¼ˆmail_meta.json ç¼ºå¤±æˆ–é”®ä¸ºç©ºï¼‰")

    # ---------- å¯¹é½ ----------
    for c in sh[cfg["center_col_letter"]]:
        c.alignment = Alignment(horizontal="center", vertical="center")
    sh[cfg["center_title_cell"]].alignment = Alignment(horizontal="left", vertical="center")

    # ---------- åˆå¹¶ä¸è¡¨å¤´ ----------
    for a, b in cfg["merge_ranges"]:
        sh.merge_cells(f"{a}:{b}")
    for addr, val in cfg["set_value_cells"].items():
        sh[addr] = val

    for i, title in enumerate(cfg["headers_K_row4"]):  # K4 èµ·è¡¨å¤´+å¡«å……
        col_letter = chr(ord("K") + i)
        cell = sh[f"{col_letter}4"]
        cell.value = title
        cell.fill = PatternFill(start_color=cfg["header_fill_color"],
                                end_color=cfg["header_fill_color"], fill_type="solid")

    # ---------- æå–ç¼–å·å†™å…¥Cåˆ— ----------
    for row in sh.iter_rows(min_row=2, max_row=last_empty_row - 1, max_col=3):
        raw = str(row[1].value).strip() if row[1].value else ""
        m = re.search(r"\d+", raw)
        if m:
            row[2].value = m.group()[-5:].zfill(5)
    sh[cfg["number_header_cell"]] = "ç¼–å·"

    # ---------- ä»â€œç¬¬ä¸€é¡µâ€ç­›é€‰ç”Ÿæˆâ€œç¬¬ä¸€é¡µå‰¯æœ¬â€ ----------
    if cfg["src_sheet_for_copy"] in wb.sheetnames:
        s1 = wb[cfg["src_sheet_for_copy"]]
        header = [c.value for c in s1[1]]
        if cfg["warehouse_col_name"] in header:
            idx = header.index(cfg["warehouse_col_name"])
            rows = [r for r in s1.iter_rows(min_row=2, values_only=True) if r[idx] == cfg["warehouse_keep_value"]]
            s_copy = wb.create_sheet(cfg["copy_sheet_name"])
            s_copy.append(header)
            for r in rows:
                s_copy.append(r)

    # ---------- ç”Ÿæˆâ€œå®¶é‡Œåº“å­˜â€ ----------
    if cfg["copy_sheet_name"] in wb.sheetnames:
        s_copy = wb[cfg["copy_sheet_name"]]
        header = [c.value for c in s_copy[1]]
        if cfg["home_name_col"] in header:
            name_idx = header.index(cfg["home_name_col"])
            qty_idx = header.index(cfg["home_qty_col"]) if cfg["home_qty_col"] in header else None

            s_home = wb.create_sheet(cfg["home_sheet_name"])
            s_home.append(cfg["home_cols"])

            for row in s_copy.iter_rows(min_row=2, values_only=True):
                name = str(row[name_idx]).strip() if row[name_idx] else ""
                code5 = name[:5] if not all('\u4e00' <= ch <= '\u9fa5' for ch in name[:5]) else ""
                qty = row[qty_idx] if qty_idx is not None else None
                if isinstance(qty, str):
                    qty = float(qty) if qty.replace(".", "", 1).isdigit() else None
                s_home.append([code5, name, qty])

            for r in s_home.iter_rows(min_row=2, max_row=s_home.max_row, min_col=3, max_col=3):
                for c in r:
                    if c.value is not None:
                        c.number_format = "#,##0.00"

    # ---------- å›å¡«åˆ°åº“å­˜è¡¨ Måˆ—ï¼ˆæ‰“å°æ¯æ¡åŒ¹é…ï¼‰ ----------
    if cfg["home_sheet_name"] in wb.sheetnames:
        s_home = wb[cfg["home_sheet_name"]]
        tgt_col = cfg["backfill_target_col_index"]  # 13 = M

        # æ„å»ºæ˜ å°„ï¼škey -> (è¯¥è¡Œçš„ Cell åˆ—è¡¨, è¡Œå·)
        map4, map5 = {}, {}
        for r in sh.iter_rows(min_row=2, max_row=last_empty_row - 1, max_col=tgt_col):
            cval = r[2].value  # Cåˆ—ï¼ˆç¼–å·ï¼‰
            if not cval:
                continue
            s = str(cval)
            # å4ä½æ˜ å°„
            if len(s) >= 4:
                map4[s[-4:]] = (r, r[0].row)
            # 5ä½æ˜ å°„ï¼ˆä¸åŸé€»è¾‘ä¸€è‡´ï¼šzfill(5)ï¼‰
            map5[s.zfill(5)] = (r, r[0].row)

        cnt_4 = cnt_5 = cnt_miss = 0

        for idx, row in enumerate(s_home.iter_rows(min_row=2, values_only=True), start=2):
            raw = str(row[0]).strip() if row[0] else ""
            name = row[1]
            qty = row[2]

            if re.fullmatch(cfg["regex_4digit_dash"], raw):
                k = raw[:4]
                if k in map4:
                    cells, rownum = map4[k]
                    c_val = cells[2].value
                    sh.cell(row=rownum, column=tgt_col).value = qty
                    print(f"âœ… å›å¡«(å4ä½åŒ¹é…) æºè¡Œ{idx} [{raw} | {name}] æ•°é‡={qty} â†’ ç›®æ ‡è¡Œ{rownum} (C={c_val}) â†’ M{rownum}")
                    cnt_4 += 1
                else:
                    print(f"â” æœªåŒ¹é…(å4ä½) æºè¡Œ{idx} [{raw} | {name}]")
                    cnt_miss += 1

            elif re.fullmatch(cfg["regex_5digit"], raw):
                k = raw.zfill(5)
                if k in map5:
                    cells, rownum = map5[k]
                    c_val = cells[2].value
                    sh.cell(row=rownum, column=tgt_col).value = qty
                    print(f"âœ… å›å¡«(5ä½åŒ¹é…)  æºè¡Œ{idx} [{raw} | {name}] æ•°é‡={qty} â†’ ç›®æ ‡è¡Œ{rownum} (C={c_val}) â†’ M{rownum}")
                    cnt_5 += 1
                else:
                    print(f"â” æœªåŒ¹é…(5ä½)   æºè¡Œ{idx} [{raw} | {name}]")
                    cnt_miss += 1
            else:
                # ç¼–ç ä¸ç¬¦åˆä¸¤ç§è§„åˆ™
                print(f"â­ï¸ è·³è¿‡(æ ¼å¼ä¸ç¬¦) æºè¡Œ{idx} [{raw} | {name}]")
                cnt_miss += 1

        print(f"ğŸ“Š å›å¡«æ±‡æ€»ï¼šå4ä½åŒ¹é… {cnt_4} æ¡ï¼Œ5ä½åŒ¹é… {cnt_5} æ¡ï¼Œæœªå‘½ä¸­/è·³è¿‡ {cnt_miss} æ¡ã€‚")

    # ---------- ä¼šè®¡æ ¼å¼ä¸å³å¯¹é½ï¼ˆG~Qï¼‰ ----------
    c1, c2 = cfg["acc_fmt_cols"]
    for col in range(c1, c2 + 1):
        letter = get_column_letter(col)
        for r in range(cfg["acc_fmt_row_from"], last_empty_row + 1):
            cell = sh[f"{letter}{r}"]
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = cfg["acc_number_format"]

    # ---------- æŒ‡å®šåˆ—å·¦å¯¹é½ï¼ˆå¦‚ T åˆ—ï¼‰ ----------
    for col_letter in cfg["left_align_cols"]:
        for cell in sh[col_letter]:
            cell.alignment = Alignment(horizontal="left")

    # ---------- åˆ—å®½ä¸éšè— ----------
    for col, w in cfg["column_widths"].items():
        sh.column_dimensions[col].width = w + 0.6
    for col in cfg["hidden_columns"]:
        sh.column_dimensions[col].hidden = True

    # ---------- è§†å›¾ï¼šå†»ç»“/ç¼©æ”¾/éšè—è¡Œ ----------
    sh.row_dimensions[1].height = cfg["row1_height"]
    sh.freeze_panes = cfg["freeze_panes"]
    sh.row_dimensions[2].outlineLevel = 1
    sh.row_dimensions[2].hidden = bool(cfg["hide_row2"])
    sh.sheet_properties.outlinePr.summaryBelow = True
    sh.sheet_view.zoomScale = cfg["zoom_scale"]

    # ---------- èšç„¦ï¼šæ¿€æ´»ç›®æ ‡è¡¨ + é€‰ä¸­ focus_cell ----------
    wb.active = wb.sheetnames.index(sh.title)
    sh.sheet_view.selection = [Selection(activeCell=cfg["focus_cell"], sqref=cfg["focus_cell"])]

    # ---------- ä¿å­˜ ----------
    wb.save(latest_file)
    wb.close()
    print(f"ğŸ‰ å·²å®Œæˆå¤„ç†å¹¶ä¿å­˜: {latest_file}")

if __name__ == "__main__":
    main(CONFIG)
