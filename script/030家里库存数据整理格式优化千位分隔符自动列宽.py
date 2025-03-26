import os
import sys
import re
from openpyxl import load_workbook
import xlwings as xw
import glob
from openpyxl.styles import PatternFill, Font, Border, Alignment

# ================================
# 📂 路径配置（支持主程序传参）
# ================================
default_folder_path = r'C:\Users\ishel\Desktop\当日库存情况'

# 通过 sys.argv 传递路径参数，兼容主控程序调用
if len(sys.argv) >= 2:
    folder_path = sys.argv[1]
    print(f"✅ 使用传入路径: {folder_path}")
else:
    folder_path = default_folder_path
    print(f"⚠️ 未传入路径，使用默认路径: {folder_path}")

# 判断路径是否存在
if not os.path.exists(folder_path):
    print(f"❌ 文件夹路径不存在: {folder_path}")
    sys.exit(1)

# === 查找路径下包含"总库存"的Excel文件 ===
files = glob.glob(os.path.join(folder_path, "*总库存*.xlsx"))

if not files:
    print("没有找到含有'总库存'的Excel文件！")
else:
    # 获取最新的“总库存”文件
    latest_file = max(files, key=os.path.getmtime)
    print(f"找到文件：{latest_file}")

    # === 打开该Excel文件并定位到“库存表” ===
    merged_wb = load_workbook(latest_file)
    if "库存表" not in merged_wb.sheetnames:
        print("⚠️ 找不到 '库存表' 工作表！")
        exit()

    sheet_kc = merged_wb["库存表"]
    print("✅ 已打开 '库存表' 工作表")

    # === 1. 解除合并单元格（排除第一行） ===
    merged_cells_ranges = list(sheet_kc.merged_cells.ranges)
    for merged_range in merged_cells_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        if min_row == 1 or max_row == 1:
            print(f"跳过合并区域（包含第一行）：{merged_range}")
            continue
        sheet_kc.unmerge_cells(str(merged_range))
        print(f"已解除合并单元格：{merged_range}")

    print("✅ 已解除库存表中除第一行以外的所有合并单元格")

    # === 新增功能：重新合并 H3:J3 区域 ===
    sheet_kc.merge_cells('H3:J3')
    print("✅ 已重新合并单元格：H3:J3")

    # === 2. 插入列 ===
    sheet_kc.insert_cols(10, 7)  # 第9列后插入7列（插入J:N列）
    sheet_kc.insert_cols(3, 1)  # 第2列后插入1列（插入C列）
    print("✅ 已插入列，C列和J:N列")

    # === 3. 提取第2列数据的后4位，写入第3列 ===
    for row in sheet_kc.iter_rows(min_row=2, max_col=3):
        cell_value = str(row[1].value).strip() if row[1].value else ""
        match = re.search(r'\d+', cell_value)
        if match:
            last_four_digits = match.group()[-4:].zfill(4)
            row[2].value = last_four_digits
    print("✅ 已提取第2列数字，写入第3列")

    # === 在C4单元格填入标题“编号” ===
    sheet_kc["C4"] = "编号"
    print("✅ 已在 C4 单元格填入标题“编号”")

    # === 4. 设置K4及其右边单元格的标题及格式 ===
    header_titles = ["外应存", "最小发货", "家里库存", "家应存", "排产", "月计划", "月已发总量"]
    j4_cell = sheet_kc["J4"]

    for i, title in enumerate(header_titles):
        col_letter = chr(ord('K') + i)  # K列开始
        cell = sheet_kc[f"{col_letter}4"]

        cell.value = title  # 设置标题

        # 假设 j4_cell 是你要复制样式的单元格
        cell.font = Font(name=j4_cell.font.name, size=j4_cell.font.size, bold=j4_cell.font.bold,
                         italic=j4_cell.font.italic, color=j4_cell.font.color)
        cell.fill = PatternFill(start_color=j4_cell.fill.start_color, end_color=j4_cell.fill.end_color,
                                fill_type=j4_cell.fill.fill_type)
        cell.border = Border(left=j4_cell.border.left, right=j4_cell.border.right, top=j4_cell.border.top,
                             bottom=j4_cell.border.bottom)
        cell.alignment = Alignment(horizontal=j4_cell.alignment.horizontal, vertical=j4_cell.alignment.vertical)

    # 填充 K4:O4 区域的背景颜色为 RGB(193, 137, 247)
    for col in range(11, 18):  # K4:O4 区域
        sheet_kc.cell(row=4, column=col).fill = PatternFill(start_color="C187F7", end_color="C187F7", fill_type="solid")

    print(f"✅ 已设置 K4 及其右边标题，并复制格式及颜色填充：{header_titles}")

    # === 5. 操作“第一页_副本” ===
    new_sheet_name = "第一页"
    if new_sheet_name in merged_wb.sheetnames:
        sheet_first = merged_wb[new_sheet_name]

        # 获取第一行作为表头
        header_row = [cell.value for cell in sheet_first[1]]
        print(f"✅ 表头: {header_row}")

        # 找到“仓库”所在的列索引
        warehouse_col_index = None
        for idx, header in enumerate(header_row):
            if header == "仓库":
                warehouse_col_index = idx + 1  # +1 因为 openpyxl 是从1开始计数
                break

        if warehouse_col_index is None:
            print("⚠️ 找不到 '仓库' 列！")
            exit()

        # 筛选“仓库”列为“成本库”的数据
        data_to_copy = []
        for row in sheet_first.iter_rows(min_row=2, max_row=sheet_first.max_row):  # 从第二行开始，排除表头
            warehouse_value = str(row[warehouse_col_index - 1].value).strip() if row[
                warehouse_col_index - 1].value else ""
            if warehouse_value == "成品库":  # 筛选“仓库”列为“成品库”的行
                data_row = [cell.value for cell in row]
                data_to_copy.append(data_row)

        if not data_to_copy:
            print("⚠️ 未找到 '成品库' 数据！")
            exit()

        # 新建“第一页副本”工作表并填充数据
        sheet_copy = merged_wb.create_sheet("第一页副本")

        # 添加表头
        sheet_copy.append(header_row)

        # 添加筛选后的数据
        for row in data_to_copy:
            sheet_copy.append(row)

        print("✅ 已创建 '第一页副本' 工作表，并复制了筛选后的 '成本库' 数据")

        # === 6. 新建“家里库存”工作表 ===
        # 获取“第一页副本”工作表
        sheet_copy = merged_wb["第一页副本"]

        # 获取第一行作为表头
        header_row = [cell.value for cell in sheet_copy[1]]

        # 找到“存货名称”所在的列索引
        inventory_name_col_index = None
        for idx, value in enumerate(header_row):
            if value == "存货名称":
                inventory_name_col_index = idx + 1  # openpyxl 中索引从1开始
                break

        if inventory_name_col_index is None:
            print("未找到 '存货名称' 列")
            exit()

        # 提取数据（从“第一页副本”获取数据）
        extracted_data = []

        for row in sheet_copy.iter_rows(min_row=2):  # 从第二行开始，排除表头
            # 获取编号：如果前5个字符都不是汉字，则提取前4个字符
            item_name_value = str(row[inventory_name_col_index - 1].value).strip() if row[
                inventory_name_col_index - 1].value else ""

            if item_name_value:
                # 判断前5个字符是否都是汉字
                if not all('\u4e00' <= char <= '\u9fa5' for char in item_name_value[:5]):  # 如果前5个字符不全为汉字
                    # 提取前4个字符作为编号
                    four_digits = item_name_value[:4]
                else:
                    four_digits = ""

                # 获取“存货名称”和“主数量”
                description = item_name_value
                quantity = str(row[header_row.index("主数量")].value) if "主数量" in header_row else ""

                extracted_data.append([four_digits, description, quantity])

        # 新建“家里库存”工作表
        home_stock_sheet = merged_wb.create_sheet("家里库存")
        home_stock_sheet.append(["编号", "存货名称", "数量"])

        # 将筛选的数据添加到“家里库存”工作表中
        for data in extracted_data:
            home_stock_sheet.append(data)

        print("✅ 已生成 '家里库存' 工作表")

    # === 6. 比较“库存表”第三列与“家里库存”编号，结果写入第13列（M列） ===
    if 'home_stock_sheet' in locals():
        home_stock_dict = {str(row[0]).zfill(4): row[2] for row in home_stock_sheet.iter_rows(min_row=2, values_only=True)}

        for row in sheet_kc.iter_rows(min_row=2, max_col=13):
            third_col_value = str(row[2].value).strip() if row[2].value else ""
            if third_col_value in home_stock_dict:
                row[12].value = home_stock_dict[third_col_value]  # 13列是M列
        print("✅ 库存表第3列和家里库存匹配，主数量已填入第13列（M列）")
    else:
        print("⚠️ '家里库存' 工作表未创建，跳过库存比较")

    # === 保存文件 ===
    # 直接使用原文件路径来保存
    merged_wb.save(latest_file)  # 使用 latest_file 保存
    merged_wb.close()
    print(f"✅ 文件已保存到原文件：{latest_file}")

    # ===================================================================
    # === xlwings 格式化（列宽、千位分隔符） ===
    # 打开 Excel 应用（可选：设置 visible=False）
    app = xw.App(visible=False)
    wb = app.books.open(latest_file)  # 打开原始文件

    # 获取“库存表”和“家里库存”工作表
    inventory_sheet = wb.sheets['库存表']
    home_inventory_sheet = wb.sheets['家里库存']
    page_one_sheet = wb.sheets['第一页']
    page_one_copy_sheet = wb.sheets['第一页副本']

    # 自动调整列宽
    inventory_sheet.autofit()
    home_inventory_sheet.autofit()
    page_one_sheet.autofit()
    page_one_copy_sheet.autofit()

    # 获取库存表、家里库存、第一页、第一页副本的已用范围
    used_range_inventory = inventory_sheet.used_range
    used_range_home_inventory = home_inventory_sheet.used_range
    used_range_page_one = page_one_sheet.used_range
    used_range_page_one_copy = page_one_copy_sheet.used_range

    # 应用千位分隔符格式
    inventory_sheet.range(used_range_inventory.address).number_format = '#,##0'
    home_inventory_sheet.range(used_range_home_inventory.address).number_format = '#,##0'
    page_one_sheet.range(used_range_page_one.address).number_format = '#,##0'
    page_one_copy_sheet.range(used_range_page_one_copy.address).number_format = '#,##0'

    print("✅ 千位分隔符格式已应用于“库存表”、“家里库存”、“第一页”和“第一页副本”，列宽已自动调整！")

    # 保存并关闭
    wb.save()  # 保存到原文件
    wb.close()
    app.quit()

print("🎉 全部处理完成！")
