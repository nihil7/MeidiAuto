import os
import sys
import glob
import openpyxl
from collections import defaultdict

# ================================
# 📂 1. 配置：确定库存文件夹路径
# ================================
# 默认目录：若在 GitHub Actions 中运行，使用 GITHUB_WORKSPACE；否则使用当前目录
default_inventory_folder = os.path.join(os.getenv("GITHUB_WORKSPACE", os.getcwd()), "data")

# 若用户通过命令行传入路径参数，则使用该路径
inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
print(f"📂 当前使用的文件夹路径: {inventory_folder}")

# 判断路径是否存在
if not os.path.exists(inventory_folder):
    print(f"❌ 目录不存在: {inventory_folder}")
    sys.exit(1)

# 匹配以“总库存”开头的 Excel 文件
files = glob.glob(os.path.join(inventory_folder, '总库存*.xlsx'))
if not files:
    print("❌ 没有找到符合条件的 Excel 文件！")
    sys.exit(1)

# 取第一个匹配文件作为处理目标
inventory_file = files[0]
print(f"✅ 找到文件：{inventory_file}")

# ================================
# 📖 2. 打开 Excel 文件并读取目标工作表
# ================================
try:
    wb_inventory = openpyxl.load_workbook(inventory_file)
except Exception as e:
    print(f"❌ 无法打开 Excel 文件: {e}")
    sys.exit(1)

# 指定要读取的明细工作表
sheet_name_detail = '出入库明细表'
if sheet_name_detail not in wb_inventory.sheetnames:
    print(f"❌ 工作表 '{sheet_name_detail}' 不存在！")
    sys.exit(1)

sheet_detail = wb_inventory[sheet_name_detail]

# ================================
# 🧾 3. 提取表头并建立列索引映射
# ================================
header_row_index = 3  # 表头所在行为第4行（从1开始计数）
headers = [cell.value for cell in sheet_detail[header_row_index]]
print(f"✅ 表头内容：{headers}")

# 将列名映射到索引（从1开始，符合 openpyxl 要求）
col_idx = {header: idx + 1 for idx, header in enumerate(headers)}

# 检查必要字段是否存在
required_columns = ['库存变动类别', '美的编码', '本期收入', '本期发出', '出入库日期']
for col in required_columns:
    if col not in col_idx:
        print(f"❌ 缺少必要列：{col}")
        sys.exit(1)

# ================================
# 🔢 4. 分类汇总：统计每个编码的出入库数据
# ================================
summary_data = defaultdict(lambda: {'入库': 0, '出库': 0})
other_records = []

# 从数据行开始逐行读取
for row in sheet_detail.iter_rows(min_row=header_row_index + 1, values_only=True):
    try:
        变动类别 = row[col_idx['库存变动类别'] - 1]
        美的编码 = row[col_idx['美的编码'] - 1]
        本期收入 = row[col_idx['本期收入'] - 1] or 0
        本期发出 = row[col_idx['本期发出'] - 1] or 0
        出入库日期 = row[col_idx['出入库日期'] - 1]

        if 变动类别 == '入库':
            summary_data[美的编码]['入库'] += 本期收入
        elif 变动类别 == '出库':
            summary_data[美的编码]['出库'] += 本期发出
        else:
            other_records.append(row)
    except Exception as e:
        print(f"⚠️ 读取行数据失败: {e}")
# ================================
# 📄 5. 创建“出入库汇总和其他变动”工作表
# ================================
sheet_name_combined = '出入库汇总和其他变动'
if sheet_name_combined in wb_inventory.sheetnames:
    del wb_inventory[sheet_name_combined]
sheet_combined = wb_inventory.create_sheet(sheet_name_combined)

# 写入汇总数据标题行
sheet_combined.append(['美的编码', '本期收入（入库）', '本期发出（出库）'])

# 写入每个编码的入库/出库总量
for 编码, data in summary_data.items():
    sheet_combined.append([编码, data['入库'], data['出库']])

# 分隔空行后写入其他变动记录（保留原始字段结构）
sheet_combined.append([])
sheet_combined.append(['录入日期', '客户子库', '单号', '美的编码', '物料品名', '单位', '仓库',
                       '库存变动类别', '本期收入', '本期发出', '条形码', '备注', '代编码', '出入库日期'])
for record in other_records:
    sheet_combined.append(record)

# ================================
# 📐 6. 自动调整列宽
# ================================
for col in sheet_combined.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    sheet_combined.column_dimensions[column].width = max_length + 2

# ================================
# 🔁 7. 将汇总的“出库”和“入库”数据写回库存表
# ================================
if '库存表' in wb_inventory.sheetnames:
    sheet_inventory = wb_inventory['库存表']

    # 提取编码列（第1列）和“入库”（第2列）、“出库”（第3列）列
    summary_first_col = [row[0] for row in sheet_combined.iter_rows(min_row=2, values_only=True)]
    summary_second_col = [row[1] for row in sheet_combined.iter_rows(min_row=2, values_only=True)]
    summary_third_col = [row[2] for row in sheet_combined.iter_rows(min_row=2, values_only=True)]

    # 获取库存表第2列（用于匹配编码）
    inventory_second_col = [row[1] for row in sheet_inventory.iter_rows(
        min_row=2, max_row=sheet_inventory.max_row, values_only=True)]

    # 写入“出库”到第18列
    for idx, inventory_value in enumerate(inventory_second_col):
        if inventory_value in summary_first_col:
            summary_index = summary_first_col.index(inventory_value)
            sheet_inventory.cell(row=idx + 2, column=18, value=summary_third_col[summary_index])

    # ✅ 添加在上面“写入出库”之后：写入“入库”到第19列
    for idx, inventory_value in enumerate(inventory_second_col):
        if inventory_value in summary_first_col:
            summary_index = summary_first_col.index(inventory_value)
            sheet_inventory.cell(row=idx + 2, column=19, value=summary_second_col[summary_index])

# ================================
# 💾 8. 保存 Excel 文件
# ================================
try:
    wb_inventory.save(inventory_file)
    print(f"✅ 完成！文件已保存: {inventory_file}")
except Exception as e:
    print(f"❌ 无法保存 Excel 文件: {e}")
    sys.exit(1)
