import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# =======================
# 配置区（按需改这里）
# =======================
DEFAULT_INV_DIR = os.path.join(os.getcwd(), "data")  # “总库存*.xlsx”所在目录
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
DEMAND_XLSX = "list.xlsx"
DEMAND_SHEET = "2503"
INV_SHEET = "库存表"

START_ROW = 5                 # 数据起始行（库存表）
WRITE_MAP = {11: "B", 14: "C", 16: "D", 20: "E"}  # K,N,P,T
ALIGN_COL_RANGE = (11, 20)    # K~T（对齐范围）
BORDER_ROWS = (4, 53)         # 行 4~53（含端点）边框
BORDER_COLS = (11, 20)        # 列 K(11)~T(20)（含端点）
FONT7_COLS = [11, 14]         # K、N 列设 7 号字
FONT7_ROWS = (4, 54)          # 行 4~54（含端点）

# =======================
# 路径与文件
# =======================
inv_dir = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_INV_DIR
if not os.path.exists(inv_dir):
    print(f"❌ 库存目录不存在: {inv_dir}"); sys.exit(1)

demand_file = os.path.join(DATA_DIR, DEMAND_XLSX)
if not os.path.exists(demand_file):
    print(f"❌ 需求文件不存在: {demand_file}"); sys.exit(1)

inventory_file = None
for f in os.listdir(inv_dir):
    if f.endswith(".xlsx") and "总库存" in f:
        inventory_file = os.path.join(inv_dir, f); break
if not inventory_file:
    print("❌ 未找到包含“总库存”的文件"); sys.exit(1)

# =======================
# 打开工作簿
# =======================
wb_demand = openpyxl.load_workbook(demand_file, data_only=True)
if DEMAND_SHEET not in wb_demand.sheetnames:
    print(f"❌ 需求缺少工作表: {DEMAND_SHEET}"); sys.exit(1)
sheet_demand = wb_demand[DEMAND_SHEET]

wb_inventory = openpyxl.load_workbook(inventory_file)
if INV_SHEET not in wb_inventory.sheetnames:
    print(f"❌ 库存缺少工作表: {INV_SHEET}"); sys.exit(1)
sheet_inventory = wb_inventory[INV_SHEET]

# =======================
# 构建映射：编码 → (B,C,D,E)
# =======================
demand_data = {}
for a, b, c, d, e in sheet_demand.iter_rows(min_row=2, max_col=5, values_only=True):
    if a is None: continue
    key = str(a).strip()
    if not key: continue
    demand_data[key] = (b, c, d, e)

# =======================
# 写入：K/N/P/T
# =======================
updated = 0
for row in sheet_inventory.iter_rows(min_row=START_ROW, max_col=20):
    code = row[2].value  # C列
    if code is None: continue
    code = str(code).strip()
    vals = demand_data.get(code)
    if not vals: continue
    r = row[0].row
    sheet_inventory.cell(row=r, column=11, value=vals[0])  # K ← B
    sheet_inventory.cell(row=r, column=14, value=vals[1])  # N ← C
    sheet_inventory.cell(row=r, column=16, value=vals[2])  # P ← D
    sheet_inventory.cell(row=r, column=20, value=vals[3])  # T ← E
    updated += 1

# =======================
# 对齐（K~T，右对齐）
# =======================
def align_range(sheet, start_row, c1, c2, h='right'):
    for col in range(c1, c2 + 1):
        for r in sheet.iter_rows(min_row=start_row, min_col=col, max_col=col):
            for cell in r:
                cell.alignment = Alignment(horizontal=h)

align_range(sheet_inventory, START_ROW, *ALIGN_COL_RANGE)

# >>> 新增：T 列左对齐 <<<
for r in sheet_inventory.iter_rows(min_row=START_ROW, min_col=20, max_col=20):
    for cell in r:
        cell.alignment = Alignment(horizontal='left')


# =======================
# 边框与字体
# =======================
thin = Border(top=Side(style="thin"), left=Side(style="thin"),
              right=Side(style="thin"), bottom=Side(style="thin"))

r1, r2 = BORDER_ROWS
c1, c2 = BORDER_COLS

# 边框 + 基础10号字
for rows in sheet_inventory.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
    for cell in rows:
        cell.border = thin
        cell.font = Font(size=10)

# 指定列改 7 号
fr1, fr2 = FONT7_ROWS
for col in FONT7_COLS:
    for rows in sheet_inventory.iter_rows(min_row=fr1, max_row=fr2, min_col=col, max_col=col):
        for cell in rows:
            cell.font = Font(size=7)

# 中文缩小为 5 号（K、N 从第 5 行起）
def has_cn(s): return bool(re.search(r'[\u4e00-\u9fff]', str(s)))
for col in [11, 14]:
    for rows in sheet_inventory.iter_rows(min_row=START_ROW, min_col=col, max_col=col):
        for cell in rows:
            if cell.value is not None and has_cn(cell.value):
                cell.font = Font(size=5)

# =======================
# 保存
# =======================
try:
    wb_inventory.save(inventory_file)
    print(f"✅ 更新 {updated} 行 | 文件: {os.path.basename(inventory_file)}")
except Exception as e:
    print(f"❌ 保存失败: {e}")
