import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ================================
# 📂 文件路径配置
# ================================
default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data"))
inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
print(f"📂 使用路径: {inventory_folder}")

if not os.path.exists(inventory_folder):
    print(f"❌ 路径不存在: {inventory_folder}")
    sys.exit(1)

# ================================
# 1. 查找文件
# ================================
pattern = os.path.join(inventory_folder, '总库存*.xlsx')
valid_files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith('~$')]

if not valid_files:
    print("❌ 没有找到符合条件的文件！")
    sys.exit(1)

inventory_file = valid_files[0]
print(f"✅ 发现库存文件: {inventory_file}")

try:
    wb_inventory = openpyxl.load_workbook(inventory_file)
    sheet_name = "库存表"
    if sheet_name not in wb_inventory.sheetnames:
        print(f"❌ 未找到工作表: {sheet_name}")
        sys.exit(1)
    sheet = wb_inventory[sheet_name]
    print(f"✅ 成功读取工作表: {sheet_name}")

    # ================================
    # 🔍 查找B列第一个空单元格所在行号
    # ================================
    col_B = sheet["B"]
    max_row = sheet.max_row
    last_empty_row = max_row + 1
    for row in range(4, max_row + 1):
        if sheet[f"B{row}"].value is None:
            last_empty_row = row
            break
    print(f"⚡ 发现 B 列第一个空单元格所在行: {last_empty_row}")

    # ================================
    # 读取表头
    # ================================
    headers = {}
    for cell in sheet[4]:
        if cell.value:
            key = cell.value.strip()
            if key not in headers:
                headers[key] = cell.column

    required_columns = [
        "外应存", "家应存", "家里库存", "库存",
        "外仓出库总量", "最小发货", "排产", "月计划", "月计划缺口"
    ]
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        print(f"❌ 缺少必要列: {missing_columns}")
        sys.exit(1)

    def col_letter(col_num):
        return openpyxl.utils.get_column_letter(col_num)

    col_external = headers["外应存"]
    col_home = headers["家应存"]
    col_stock = headers["家里库存"]
    col_total_stock = headers["库存"]
    col_external_ship = headers["外仓出库总量"]
    col_min_ship = headers["最小发货"]
    col_production = headers["排产"]
    col_plan = headers["月计划"]
    col_gap = headers["月计划缺口"]
    col_ref = 10  # J列

    print("✅ 表头索引解析完成")

    gray_font = Font(color="D8D8D8")
    default_font = Font(color="000000")

    def safe_float(value):
        try:
            return float(value) if value else 0
        except ValueError:
            return 0

    DEBUG_PRINT = True
    DEBUG_ROWS = []

    for row_idx in range(5, last_empty_row):  # ✅ 限制处理行范围
        external_stock = safe_float(sheet[f"{col_letter(col_external)}{row_idx}"].value)
        home_stock = safe_float(sheet[f"{col_letter(col_home)}{row_idx}"].value)
        stock_at_home = safe_float(sheet[f"{col_letter(col_stock)}{row_idx}"].value)
        total_stock = safe_float(sheet[f"{col_letter(col_total_stock)}{row_idx}"].value)
        external_shipped = safe_float(sheet[f"{col_letter(col_external_ship)}{row_idx}"].value)
        month_plan = safe_float(sheet[f"{col_letter(col_plan)}{row_idx}"].value)
        ref_value = safe_float(sheet[f"{col_letter(col_ref)}{row_idx}"].value)

        min_ship_result = external_stock - ref_value
        production_result = home_stock + external_stock - ref_value - stock_at_home
        gap_result = month_plan - stock_at_home - total_stock - external_shipped

        if DEBUG_PRINT and (not DEBUG_ROWS or row_idx in DEBUG_ROWS):
            print(f"🔍 行 {row_idx} | 月计划: {month_plan:.1f}, 家里库存: {stock_at_home:.1f}, "
                  f"库存: {total_stock:.1f}, 外仓出库: {external_shipped:.1f} → 缺口: {gap_result:.1f}")

        sheet[f"{col_letter(col_min_ship)}{row_idx}"].value = min_ship_result
        sheet[f"{col_letter(col_production)}{row_idx}"].value = production_result
        sheet[f"{col_letter(col_gap)}{row_idx}"].value = gap_result

        sheet[f"{col_letter(col_min_ship)}{row_idx}"].font = gray_font if min_ship_result <= 0 else default_font
        sheet[f"{col_letter(col_production)}{row_idx}"].font = gray_font if production_result <= 0 else default_font
        sheet[f"{col_letter(col_gap)}{row_idx}"].font = gray_font if gap_result <= 0 else default_font

    print("✅ 公式计算完成")

    # ================================
    # 写入 G～U 列合计公式（非负数值求和）
    # ================================
    print(f"✅ 计算求和的目标行: {last_empty_row}")
    for col in range(7, 22):  # G~U
        col_letter = get_column_letter(col)
        start_row = 5
        end_row = last_empty_row - 1
        range_expr = f"{col_letter}{start_row}:{col_letter}{end_row}"

        sum_formula = (
            f'=SUMPRODUCT(--(ISNUMBER(--{range_expr})), '
            f'--(--{range_expr}>=0), '
            f'--(--{range_expr}))'
        )
        cell_addr = f"{col_letter}{last_empty_row}"
        sum_cell = sheet[cell_addr]

        if sum_cell.value is not None:
            print(f"⚠️  原有值将被覆盖 → {cell_addr} 原值: {sum_cell.value}")
        else:
            print(f"🆕 即将写入 → {cell_addr}")

        sum_cell.value = sum_formula
        print(f"✅ 已写入公式至 {cell_addr}: {sum_formula}")

    # ================================
    # 保存Excel文件
    # ================================
    wb_inventory.save(inventory_file)
    wb_inventory.close()
    print(f"🎉 文件已保存: {inventory_file}")

except Exception as e:
    print(f"❌ Excel 处理失败: {e}")
    sys.exit(1)
