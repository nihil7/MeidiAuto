import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font

# ================================
# 📂 文件路径配置（支持主程序传参）
# ================================
default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data", "mail"))

inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
print(f"📂 使用路径: {inventory_folder}")

if not os.path.exists(inventory_folder):
    print(f"❌ 路径不存在: {inventory_folder}")
    sys.exit(1)

# ================================
# 1. 查找文件（过滤临时文件）
# ================================
pattern = os.path.join(inventory_folder, '总库存*.xlsx')
valid_files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith('~$')]

if not valid_files:
    print("❌ 没有找到符合条件的文件！")
    sys.exit(1)

inventory_file = valid_files[0]
print(f"✅ 发现库存文件: {inventory_file}")

# ================================
# 2. 读取Excel文件，获取工作表
# ================================
try:
    wb_inventory = openpyxl.load_workbook(inventory_file)
    sheet_name = "库存表"

    if sheet_name not in wb_inventory.sheetnames:
        print(f"❌ 未找到工作表: {sheet_name}")
        sys.exit(1)

    sheet = wb_inventory[sheet_name]
    print(f"✅ 成功读取工作表: {sheet_name}")

    # ================================
    # 3. 读取表头并检查必要列
    # ================================
    headers = {cell.value.strip(): cell.column for cell in sheet[4] if cell.value}
    required_columns = ["外应存", "家应存", "家里库存", "最小发货", "排产"]
    missing_columns = [col for col in required_columns if col not in headers]

    if missing_columns:
        print(f"❌ 缺少必要列: {missing_columns}")
        sys.exit(1)


    # 获取列索引
    def col_letter(col_num):
        return openpyxl.utils.get_column_letter(col_num)


    col_external = headers["外应存"]
    col_home = headers["家应存"]
    col_stock = headers["家里库存"]
    col_min_ship = headers["最小发货"]
    col_production = headers["排产"]
    col_ref = 10  # 参照列

    print("✅ 表头索引解析完成")

    # ================================
    # 4. 遍历数据行，计算公式并更新颜色
    # ================================
    gray_font = Font(color="D8D8D8")
    default_font = Font(color="000000")


    def safe_float(value):
        try:
            return float(value) if value else 0
        except ValueError:
            return 0


    for row_idx in range(5, sheet.max_row + 1):
        cell_external = f"{col_letter(col_external)}{row_idx}"
        cell_home = f"{col_letter(col_home)}{row_idx}"
        cell_stock = f"{col_letter(col_stock)}{row_idx}"
        cell_ref = f"{col_letter(col_ref)}{row_idx}"

        cell_min_ship = sheet[f"{col_letter(col_min_ship)}{row_idx}"]
        cell_production = sheet[f"{col_letter(col_production)}{row_idx}"]

        external_stock = safe_float(sheet[cell_external].value)
        home_stock = safe_float(sheet[cell_home].value)
        stock_at_home = safe_float(sheet[cell_stock].value)
        ref_value = safe_float(sheet[cell_ref].value)

        min_ship_result = external_stock - ref_value
        production_result = home_stock + external_stock - ref_value - stock_at_home

        cell_min_ship.value, cell_production.value = min_ship_result, production_result
        cell_min_ship.font = gray_font if min_ship_result <= 0 else default_font
        cell_production.font = gray_font if production_result <= 0 else default_font

    print("✅ 公式计算完成，正在保存文件...")

    # ================================
    # 5. 调整列宽
    # ================================
    col_widths = {
        'B': 8, 'C': 6, 'D': 36.88, 'E': 3, 'F': 3.6,
        'G': 6.98, 'H': 6.98, 'I': 6.98, 'J': 6.98,
        'K': 4.8,'L': 8.1, 'M': 6.98, 'N': 5,
        'O': 6.98, 'P': 7.9, 'Q': 7.9
    }
    for col, width in col_widths.items():
        sheet.column_dimensions[col].width = width + 0.6

    print("✅ 列宽调整完成")

    # ================================
    # 6. 保存Excel文件
    # ================================
    wb_inventory.save(inventory_file)
    wb_inventory.close()

    print(f"🎉 文件已保存: {inventory_file}")

except Exception as e:
    print(f"❌ Excel 处理失败: {e}")
    sys.exit(1)
