# -*- coding: utf-8 -*-
"""
库存表着色脚本（覆盖保存到原文件）

✅ 本版已修复：淡红/淡紫扩展时“空单元格不着色”的问题（现在整段范围都会铺满）
✅ 支持：淡色扩展两段 A~K + M~T（可配置）
✅ 支持：按某列（默认C列）命中指定值整行跳过着色（可配置）
✅ 深色（深紫/深红/绿）只涂在 n 列（默认L列），不会被淡色覆盖
"""

import os
import sys
import glob
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# =========================================================
# ✅ 配置区（只改这里即可）
# =========================================================

# 1) 文件/工作表
DEFAULT_INVENTORY_FOLDER = Path.cwd() / "data"
FILE_GLOB_PATTERN = "总库存*.xlsx"
PICK_LATEST_FILE = True          # True=选最新修改的文件；False=按glob第一个
SHEET_NAME = "库存表"
LOAD_DATA_ONLY = False           # True=读取公式缓存值；False=读取公式本身（一般用False）

# 2) 逻辑列定义（Excel列字母）
COL_M = "J"                      # m 所在列（默认 J=第10列）
COL_N = "L"                      # n 所在列（默认 L=第12列） -> 深色（深紫/深红/绿）会涂在这一列

# 3) 跳过整行着色（按某列的值）
SKIP_COL = "C"                   # 哪一列用于判断跳过
SKIP_CODES = {"00514", "04928"}  # 命中则整行不参与着色
SKIP_CODE_DIGITS = 5             # 统一补零位数（00514）
CLEAR_FILL_ON_SKIPPED_ROW = False  # True=命中跳过时清掉该行A~T范围底色；False=不动原底色

# 4) 淡色扩展范围（仅在“深紫/深红”两种情况触发）
ROW_FILL_RANGES = [("A", "K"), ("M", "T")]  # 你要的两段：A~K + M~T

# ✅ 关键修复：是否只给“非空”单元格涂淡色
FILL_ONLY_NON_EMPTY = False      # False=空单元格也涂（整段铺满）；True=只涂有值的格子

# 5) 颜色配置（RGB 十六进制，别带 #）
COLOR_DEEP_PURPLE = "3F0065"     # 深紫：仅涂 COL_N
COLOR_DEEP_RED    = "FF0000"     # 深红：仅涂 COL_N
COLOR_GREEN       = "00FF00"     # 绿：仅涂 COL_N（不扩展淡色）
COLOR_LIGHT_PURPLE = "CCC0DA"    # 淡紫：扩展到 ROW_FILL_RANGES
COLOR_LIGHT_RED    = "E6B8B7"    # 淡红：扩展到 ROW_FILL_RANGES

# 6) 输出
VERBOSE = True                   # True=打印每行着色信息（会比较多）


# =========================================================
# 工具函数
# =========================================================

def _fill(color_hex: str) -> PatternFill:
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

def normalize_code(v, digits: int = 5) -> str:
    """把跳过列的值统一成 digits 位数字字符串，兼容 '00514' / '514' / 514 / 514.0 等"""
    if v is None:
        return ""
    s = str(v).strip()

    # 处理 514.0 / "514.0"
    try:
        if isinstance(v, (int, float)) or (s.replace(".", "", 1).isdigit() and "." in s):
            s = str(int(float(v)))
    except Exception:
        pass

    if s.isdigit():
        s = s.zfill(digits)
    return s

def safe_float(v) -> float:
    """把单元格值安全转成 float。若是空/非数字（如公式字符串），返回0"""
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def required_max_col() -> int:
    """根据配置自动计算本次至少要遍历到哪一列"""
    cols = [COL_M, COL_N, SKIP_COL]
    for a, b in ROW_FILL_RANGES:
        cols.extend([a, b])
    return max(column_index_from_string(c) for c in cols)

def pick_inventory_file(folder_path: Path) -> str:
    pattern = str(folder_path / FILE_GLOB_PATTERN)
    files = glob.glob(pattern)
    valid = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not valid:
        raise FileNotFoundError(f"没有找到符合条件的文件：{pattern}")

    if PICK_LATEST_FILE:
        valid.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return valid[0]

def iter_col_indices(a: str, b: str):
    start = column_index_from_string(a)
    end = column_index_from_string(b)
    if start > end:
        start, end = end, start
    return range(start, end + 1)

def should_fill_light(cell) -> bool:
    if not FILL_ONLY_NON_EMPTY:
        return True
    return cell.value not in (None, "")

def clear_row_fills(sheet, row_idx: int, max_col: int):
    empty_fill = PatternFill()
    for col_idx in range(1, max_col + 1):
        sheet.cell(row=row_idx, column=col_idx).fill = empty_fill

def apply_light_fill(sheet, row_idx: int, light_fill: PatternFill, exclude_cols_letters: set):
    """
    给 ROW_FILL_RANGES 指定范围铺淡色：
    - 现在允许“空单元格也涂”（由 FILL_ONLY_NON_EMPTY 控制）
    - 不再限制 sheet.max_column，确保 M~T 即使原本为空也能被涂出来
    """
    for a, b in ROW_FILL_RANGES:
        for col_idx in iter_col_indices(a, b):
            col_letter = get_column_letter(col_idx)
            if col_letter in exclude_cols_letters:
                continue
            cell = sheet.cell(row=row_idx, column=col_idx)
            if should_fill_light(cell):
                cell.fill = light_fill


# =========================================================
# 核心处理
# =========================================================

def process_inventory_data(sheet):
    deep_purple = _fill(COLOR_DEEP_PURPLE)
    deep_red = _fill(COLOR_DEEP_RED)
    green = _fill(COLOR_GREEN)

    light_purple = _fill(COLOR_LIGHT_PURPLE)
    light_red = _fill(COLOR_LIGHT_RED)

    idx_m = column_index_from_string(COL_M) - 1
    idx_n = column_index_from_string(COL_N) - 1
    idx_skip = column_index_from_string(SKIP_COL) - 1

    max_col = required_max_col()

    # 淡色扩展时，保护深色列不被覆盖
    exclude_light = {COL_N}

    # 遍历：A ~ max_col（比如到T=20列）
    for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=False):
        row_idx = row[0].row

        # 1) 跳过整行
        skip_val = normalize_code(row[idx_skip].value, SKIP_CODE_DIGITS) if idx_skip < len(row) else ""
        if skip_val in SKIP_CODES:
            if CLEAR_FILL_ON_SKIPPED_ROW:
                clear_row_fills(sheet, row_idx, max_col)
            if VERBOSE:
                print(f"行 {row_idx} → 跳过着色（{SKIP_COL}列={skip_val}）")
            continue

        # 2) 取 m / n
        m = safe_float(row[idx_m].value) if idx_m < len(row) else 0.0
        n = safe_float(row[idx_n].value) if idx_n < len(row) else 0.0

        # 只处理 n>0
        if n <= 0:
            continue

        cell_n = row[idx_n]  # COL_N 的单元格

        # 情况1：深紫 + 淡紫扩展
        if (m == 0 and n > 0) or (m < 0):
            cell_n.fill = deep_purple
            apply_light_fill(sheet, row_idx, light_purple, exclude_light)
            if VERBOSE:
                print(f"行 {row_idx} → 深紫({COL_N}) + 淡紫铺色: n={n}, m={m}")

        # 情况2：绿色（不扩展淡色）
        elif m != 0 and (n / m) < 1:
            cell_n.fill = green
            if VERBOSE:
                print(f"行 {row_idx} → 绿色({COL_N}): n={n}, m={m}")

        # 情况3：深红 + 淡红扩展
        elif m != 0 and (n / m) >= 1:
            cell_n.fill = deep_red
            apply_light_fill(sheet, row_idx, light_red, exclude_light)
            if VERBOSE:
                print(f"行 {row_idx} → 深红({COL_N}) + 淡红铺色: n={n}, m={m}")


def main(folder_path: str):
    folder = Path(folder_path).resolve()
    if not folder.exists():
        print(f"❌ 文件夹路径不存在: {folder}")
        sys.exit(1)

    inventory_file = pick_inventory_file(folder)
    print(f"✅ 找到文件：{inventory_file}")

    wb = openpyxl.load_workbook(inventory_file, data_only=LOAD_DATA_ONLY)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ 工作表不存在：{SHEET_NAME}，实际为：{wb.sheetnames}")
        sys.exit(1)

    sheet = wb[SHEET_NAME]
    process_inventory_data(sheet)

    wb.save(inventory_file)
    print(f"✅ 处理后的文件已保存（覆盖原文件）：{inventory_file}")


if __name__ == "__main__":
    # 支持外部传参路径（来自主程序）
    if len(sys.argv) >= 2:
        inventory_folder = sys.argv[1]
        print(f"✅ 使用传入路径: {inventory_folder}")
    else:
        inventory_folder = str(DEFAULT_INVENTORY_FOLDER)
        print(f"⚠️ 未传入路径，使用默认路径: {inventory_folder}")

    main(inventory_folder)
