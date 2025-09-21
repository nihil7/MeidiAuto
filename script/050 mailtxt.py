import os
import sys
import glob
import openpyxl
from datetime import datetime
import re


# ================================
# ğŸ“‚ é…ç½®æ–‡ä»¶è·¯å¾„
# ================================
def get_inventory_folder():
    default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data"))
    inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
    print(f"âœ… ä½¿ç”¨ä¼ å…¥è·¯å¾„: {inventory_folder}" if len(
        sys.argv) >= 2 else f"âš ï¸ æœªä¼ å…¥è·¯å¾„ï¼Œä½¿ç”¨é»˜è®¤è·¯å¾„: {inventory_folder}")

    if not os.path.exists(inventory_folder):
        print(f"âŒ æ–‡ä»¶å¤¹è·¯å¾„ä¸å­˜åœ¨: {inventory_folder}")
        sys.exit(1)

    print(f"ğŸ“‚ å½“å‰å·¥ä½œæ–‡ä»¶å¤¹: {inventory_folder}")
    return inventory_folder


# ================================
# 1. æŸ¥æ‰¾ Excel æ–‡ä»¶
# ================================
def find_excel_file(inventory_folder):
    files = glob.glob(os.path.join(inventory_folder, 'æ€»åº“å­˜*.xlsx'))
    valid_files = [f for f in files if not os.path.basename(f).startswith('~$')]

    if not valid_files:
        print("âŒ æ²¡æœ‰æ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„æ–‡ä»¶ï¼")
        sys.exit(1)

    print(f"âœ… æ‰¾åˆ°æ–‡ä»¶ï¼š{valid_files[0]}")
    return valid_files[0]


# ================================
# 2. è¯»å–å·¥ä½œè¡¨
# ================================
def load_worksheet(inventory_file, sheet_name="åº“å­˜è¡¨"):
    try:
        wb = openpyxl.load_workbook(inventory_file)
    except Exception as e:
        print(f"âŒ æ— æ³•æ‰“å¼€ Excel æ–‡ä»¶ï¼š{e}")
        sys.exit(1)

    if sheet_name not in wb.sheetnames:
        print(f"âŒ å·¥ä½œè¡¨â€œ{sheet_name}â€ä¸å­˜åœ¨ï¼")
        sys.exit(1)

    return wb[sheet_name]


# ================================
# ğŸ¨ é¢œè‰²åˆ¤æ–­å‡½æ•°ï¼ˆåªè¯†åˆ«å¡«å……è‰²ï¼‰
# ================================
def get_cell_fill_rgb(cell):
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            return color.rgb[-6:].upper()
    return None


def is_fill_color(cell, color_code: str):
    return get_cell_fill_rgb(cell) == color_code.upper()


# ================================
# æŸ¥æ‰¾çº¢è‰²æˆ–ç´«è‰²çš„è¡Œ
# ================================
def find_colored_rows(sheet):
    colored_rows = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=12)
        if is_fill_color(cell, "FF0000") or is_fill_color(cell, "3F0065"):
            colored_rows.append(row)
            color_hex = get_cell_fill_rgb(cell)
            print(f"âœ… ç¬¦åˆæ¡ä»¶é¢œè‰² â†’ è¡Œ: {row} RGB: #{color_hex}")
    return colored_rows


# ================================
# ğŸ“… è·å–æ—¥æœŸï¼ˆH3 ä¸ M3ï¼‰
# ================================
def _fmt_dt(v):
    """æŠŠå•å…ƒæ ¼æ—¶é—´æˆ–å­—ç¬¦ä¸²æ ¼å¼åŒ–ä¸º 'YYYY-MM-DD HH:MM:SS'ï¼›æ— æ³•è§£æå°±åŸæ ·è¿”å›/ç©ºä¸²ã€‚"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        try:
            # ISO8601 å…œåº•ï¼Œå¦‚ 2025-09-19T18:00:05 æˆ–å¸¦Z
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return s
    return ""


def get_dates(sheet):
    """
    è¿”å› (date, date2)
    - date  æ¥è‡ª H3ï¼ˆåŸæ ‡é¢˜ç”¨ï¼‰
    - date2 æ¥è‡ª M3ï¼ˆå®¶é‡Œåº“å­˜æ•°æ®ç”¨ï¼‰
    """
    date = _fmt_dt(sheet["H3"].value)
    date2 = _fmt_dt(sheet["M3"].value)
    return date, date2


# ================================
# æŸ¥æ‰¾ B åˆ—ç¬¬ä¸€ä¸ªç©ºå•å…ƒæ ¼
# ================================
def find_last_empty_row(sheet):
    for row in range(4, sheet.max_row + 1):
        if sheet[f"B{row}"].value is None:
            return row
    return sheet.max_row + 1


# ================================
# è®¡ç®—å…¬å¼å¹¶è¿”å›æ€»å’Œ
# ================================
def calculate_sum(sheet, formula):
    if isinstance(formula, (int, float)):
        return formula
    if not isinstance(formula, str):
        return 0

    match = re.match(r"^=SUM\((.+)\)$", formula)
    if match:
        cell_range = match.group(1)
        start_cell, end_cell = cell_range.split(":")
        start_row, start_col = int(start_cell[1:]), openpyxl.utils.cell.column_index_from_string(start_cell[:1])
        end_row, end_col = int(end_cell[1:]), openpyxl.utils.cell.column_index_from_string(end_cell[:1])

        total = 0
        for row in range(start_row, end_row + 1):
            value = sheet.cell(row=row, column=start_col).value
            if isinstance(value, (int, float)):
                total += value
        return total
    return 0


# ================================
# è·å–åº“å­˜åˆè®¡ä¿¡æ¯
# ================================
def prepare_summary_text(sheet, last_empty_row):
    stock_total       = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=10).value)
    monthly_plan      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=16).value)
    plan_gap_output   = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=17).value)
    monthly_sent      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=18).value)
    monthly_received  = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=19).value)
    monthly_remaining = monthly_plan - monthly_sent if monthly_plan and monthly_sent else 0

    print(f"ğŸ“Š åº“å­˜æ€»é‡: {stock_total}, æœˆè®¡åˆ’: {monthly_plan}, ç¼ºå£æ’äº§: {plan_gap_output}, å‡ºåº“: {monthly_sent}, å…¥åº“: {monthly_received}")
    return stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining


# ================================
# æ„å»ºè¾“å‡ºæ–‡æœ¬
# ================================
def construct_html_content(sheet, colored_rows, date, date2,
                           stock_total, monthly_plan, plan_gap_output,
                           monthly_sent, monthly_received, monthly_remaining):
    html = """
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            table { border-collapse: collapse; width: auto; margin-top: 10px; }
            th, td { border: 1px solid #999; padding: 6px 10px; }
            th { background-color: #f2f2f2; text-align: left; }
            td.right { text-align: right; }
            td.left { text-align: left; }
        </style>
    </head>
    <body>
    """

    # ä¸¤ä¸ªæ ‡é¢˜ï¼šH3 å¯¹åº”é‡åº†ä¿Šéƒ½ä»“å‚¨ï¼ŒM3 å¯¹åº”å®¶é‡Œåº“å­˜
    html += f"""
    <h1>{date} é‡åº†ä¿Šéƒ½ä»“å‚¨æ•°æ®</h1>
    <h1>{date2} å®¶é‡Œåº“å­˜æ•°æ®</h1>
    <h5>â€œå¤–ä»“åº“å­˜ï¼œ50%å¤–ä»“åº”å­˜æ•°é‡â€çš„ç‰©æ–™æœ‰ <strong>{len(colored_rows)}</strong> æ¬¾</h5>
    """

    html += """
    <table>
        <tr>
            <th>ç¼–å·</th>
            <th>åº“å­˜</th>
            <th>å¤–åº”å­˜</th>
            <th>å®¶é‡Œåº“å­˜</th>
        </tr>
    """
    for row in colored_rows:
        code = sheet.cell(row=row, column=3).value
        stock = sheet.cell(row=row, column=10).value
        expected = sheet.cell(row=row, column=11).value
        home_stock = sheet.cell(row=row, column=13).value

        stock_fmt = f"{stock:,.1f}" if isinstance(stock, (int, float)) else stock
        expected_fmt = f"{expected:,.1f}" if isinstance(expected, (int, float)) else expected
        home_stock_fmt = f"{home_stock:,.1f}" if isinstance(home_stock, (int, float)) else home_stock

        html += f"""
        <tr>
            <td>{code}</td>
            <td class="right">{stock_fmt}</td>
            <td class="right">{expected_fmt}</td>
            <td class="right">{home_stock_fmt}</td>
        </tr>
        """

    html += "</table>"

    html += """
    <h5>æ±‡æ€»ä¿¡æ¯</h5>
    <table>
        <tr><th>é¡¹ç›®</th><th>æ•°å€¼</th></tr>
    """

    def row(label, value):
        return f"""
        <tr>
            <td class="left">{label}</td>
            <td class="right">{value:,.1f}</td>
        </tr>
        """

    html += row("å¤–ä»“åº“å­˜æ€»é‡", stock_total)
    html += row("æœˆè®¡åˆ’", monthly_plan)
    html += row("æœˆè®¡åˆ’ç¼ºå£æ’äº§", plan_gap_output)
    html += row("å¤–ä»“å‡ºåº“æ€»é‡", monthly_sent)
    html += row("å¤–ä»“å…¥åº“æ€»é‡", monthly_received)
    html += row("æœˆé¢„ä¼°è¿˜æœ‰è¦å‘è´§", monthly_remaining)

    html += "</table>\n</body></html>"
    return html


# ================================
# ä¿å­˜ä¸º HTML æ–‡ä»¶
# ================================
def save_output_to_file(html_content, output_dir):
    output_filename = os.path.join(output_dir, "output.html")
    with open(output_filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"ğŸ“ å·²æˆåŠŸä¿å­˜ä¸º HTML æ–‡ä»¶ï¼š{output_filename}")


# ================================
# ä¸»å‡½æ•°
# ================================
def main():
    inventory_folder = get_inventory_folder()
    inventory_file = find_excel_file(inventory_folder)
    sheet = load_worksheet(inventory_file)

    colored_rows = find_colored_rows(sheet)
    date, date2 = get_dates(sheet)  # ğŸ‘ˆ åŒæ—¶æ‹¿ H3 / M3
    last_empty_row = find_last_empty_row(sheet)
    print(f"âš¡ å‘ç° B åˆ—ç¬¬ä¸€ä¸ªç©ºå•å…ƒæ ¼æ‰€åœ¨è¡Œ: {last_empty_row}")

    stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining = prepare_summary_text(sheet, last_empty_row)

    html_content = construct_html_content(
        sheet, colored_rows, date, date2,
        stock_total, monthly_plan, plan_gap_output,
        monthly_sent, monthly_received, monthly_remaining
    )

    print("\nğŸ“‹ HTML å·²ç”Ÿæˆï¼Œé¢„è§ˆå†…å®¹çœç•¥â€¦")
    save_output_to_file(html_content, inventory_folder)
    print("âœ… çº¢è‰²æˆ–ç´«è‰²å•å…ƒæ ¼æ•°é‡ï¼š", len(colored_rows))
    print("ğŸ“Œ è¡Œå·åˆ—è¡¨ï¼š", colored_rows)


if __name__ == "__main__":
    main()
