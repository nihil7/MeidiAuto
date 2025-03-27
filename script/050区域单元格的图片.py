import sys
import os
import glob
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from matplotlib.font_manager import FontProperties
from matplotlib import rcParams

# 设置中文字体（以SimHei为例）
rcParams['font.family'] = 'SimHei'

# 1. 文件路径配置
# ================================
default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data", "mail"))

# 判断是否传入路径
if len(sys.argv) >= 2:
    inventory_folder = sys.argv[1]
    print(f"✅ 使用传入路径: {inventory_folder}")
else:
    inventory_folder = default_inventory_folder
    print(f"⚠️ 未传入路径，使用默认路径: {inventory_folder}")

# 确保文件夹路径存在
if not os.path.exists(inventory_folder):
    print(f"❌ 文件夹路径不存在: {inventory_folder}")
    exit()

# 匹配文件：总库存*.xlsx
pattern = os.path.join(inventory_folder, '总库存*.xlsx')
files = glob.glob(pattern)

# 确保文件存在
if not files:
    print("❌ 没有找到符合条件的文件！")
    exit()

# 获取最新文件
latest_file = max(files, key=os.path.getctime)
print(f"✅ 找到最新的文件：{latest_file}")

# ================================
# 2. 使用 openpyxl 读取 Excel 文件并提取数据
# ================================

# 打开 Excel 文件
wb = openpyxl.load_workbook(latest_file)
ws = wb.active  # 默认选择第一个工作表

# 定义区域 A1:Q60
data = []
cell_styles = []
col_widths = []  # 用于存储列宽

# 获取 A1:Q60 区域的内容以及样式
for row in ws['A1:Q60']:  # 设置区域，只读取 60 行
    row_data = []
    row_styles = []
    for cell in row:
        row_data.append(cell.value)
        row_styles.append({
            'font_color': cell.font.color.rgb if cell.font.color else None,
            'fill_color': cell.fill.start_color.rgb if cell.fill.start_color else None,
            'border': cell.border,
            'font_name': cell.font.name,
            'font_size': cell.font.size,
            'font_bold': cell.font.bold,
            'font_italic': cell.font.italic,
            'font_underline': cell.font.underline,
        })
    data.append(row_data)
    cell_styles.append(row_styles)

# 获取列宽
for col in ws.columns:
    col_widths.append(max(len(str(cell.value)) for cell in col))

# 转换为 NumPy 数组，方便绘制图片
data_np = np.array(data)

# ================================
# 3. 使用 matplotlib 绘制表格并保存为图片
# ================================

# 创建图形和轴
fig, ax = plt.subplots(figsize=(10, 6))

# 隐藏坐标轴
ax.axis('tight')
ax.axis('off')

# 创建表格
table = ax.table(cellText=data_np, loc='center', cellLoc='center', colLabels=[cell.value for cell in ws[1]],
                 rowLabels=[f"Row {i}" for i in range(1, len(data) + 1)])  # 动态行标签长度

# 应用样式
for (i, j), cell in table.get_celld().items():
    # 确保i和j不超出cell_styles的范围
    if i < len(cell_styles) and j < len(cell_styles[i]):
        # 获取单元格的样式
        font_color = cell_styles[i][j]['font_color']
        fill_color = cell_styles[i][j]['fill_color']
        font_name = cell_styles[i][j]['font_name']
        font_size = cell_styles[i][j]['font_size']
        font_bold = cell_styles[i][j]['font_bold']
        font_italic = cell_styles[i][j]['font_italic']
        font_underline = cell_styles[i][j]['font_underline']

        # 设置字体颜色
        if font_color and font_color != '00000000':  # '00000000' 是没有颜色的情况
            if isinstance(font_color, str) and font_color.startswith('00'):
                r, g, b = [int(font_color[i:i + 2], 16) for i in (2, 4, 6)]  # 跳过前两位'00'
                # 设置字体颜色
                font_props = FontProperties(weight='bold' if font_bold else 'normal',
                                            size=font_size if font_size else 10,
                                            style='italic' if font_italic else 'normal',
                                            variant='normal' if font_underline else 'normal')  # 设置字体样式
                cell.set_text_props(color=(r / 255, g / 255, b / 255), fontproperties=font_props)

        # 设置填充颜色
        if fill_color and fill_color != '00000000':  # 同理，处理无填充颜色的情况
            if isinstance(fill_color, str) and fill_color.startswith('00'):
                r, g, b = [int(fill_color[i:i + 2], 16) for i in (2, 4, 6)]  # 跳过前两位'00'
                cell.set_facecolor((r / 255, g / 255, b / 255))

# 获取当前时间，命名图片
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
image_filename = f"美的仓储自动化_{current_time}.png"
image_filepath = os.path.join(inventory_folder, image_filename)

# 保存为高清图片，设置高分辨率（dpi=300）
plt.savefig(image_filepath, bbox_inches='tight', pad_inches=0.05, dpi=1200)
plt.close()

print(f"✅ 图片已保存：{image_filepath}")
