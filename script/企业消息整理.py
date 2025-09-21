import os
import sys
import glob
import openpyxl
from datetime import datetime
import re

# ================================
# 📂 配置
# ================================
MAX_LIST = 20  # 明细最多展示条数（可按需调整）

# ================================
# 📂 路径
# ================================
def get_inventory_folder():
    default_inventory_folder = os.path.abspath(os.path.join(os.getcwd(), "data"))
    inventory_folder = sys.argv[1] if len(sys.argv) >= 2 else default_inventory_folder
    print(f"✅ 使用传入路径: {inventory_folder}" if len(sys.argv) >= 2 else f"⚠️ 未传入路径，使用默认路径: {inventory_folder}")

    if not os.path.exists(inventory_folder):
        print(f"❌ 文件夹路径不存在: {inventory_folder}")
        sys.exit(1)

    print(f"📂 当前工作文件夹: {inventory_folder}")
    return inventory_folder

# ================================
# 1) 查找 Excel 文件
# ================================
def find_excel_file(inventory_folder):
    files = glob.glob(os.path.join(inventory_folder, '总库存*.xlsx'))
    valid_files = [f for f in files if not os.path.basename(f).startswith('~$')]
    if not valid_files:
        print("❌ 没有找到符合条件的文件！")
        sys.exit(1)
    print(f"✅ 找到文件：{valid_files[0]}")
    return valid_files[0]

# ================================
# 2) 打开工作表
# ================================
def load_worksheet(inventory_file, sheet_name="库存表"):
    try:
        wb = openpyxl.load_workbook(inventory_file, data_only=True)
    except Exception as e:
        print(f"❌ 无法打开 Excel 文件：{e}")
        sys.exit(1)

    if sheet_name not in wb.sheetnames:
        print(f"❌ 工作表“{sheet_name}”不存在！")
        sys.exit(1)

    return wb[sheet_name]

# ================================
# 🎨 颜色判断（仅识别填充色）
# ================================
def get_cell_fill_rgb(cell):
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            return color.rgb[-6:].upper()
    return None

def is_fill_color(cell, color_code: str):
    return get_cell_fill_rgb(cell) == color_code.upper()

# ================================
# 查找红色/紫色行（L列=12）
# ================================
def find_colored_rows(sheet):
    colored_rows = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=12)
        if is_fill_color(cell, "FF0000") or is_fill_color(cell, "3F0065"):
            colored_rows.append(row)
            color_hex = get_cell_fill_rgb(cell)
            print(f"✅ 符合条件颜色 → 行: {row} RGB: #{color_hex}")
    return colored_rows

# ================================
# 📅 日期（H3 与 M3）
# ================================
def _fmt_dt(v):
    """格式化为 'YYYY-MM-DD HH:MM:SS'；无法解析则原样/空串。"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return s
    return ""

def get_dates(sheet):
    """返回 (date, date2) —— H3（仓储）与 M3（家里库存）"""
    date = _fmt_dt(sheet["H3"].value)
    date2 = _fmt_dt(sheet["M3"].value)
    return date, date2

# ================================
# B列第一个空行
# ================================
def find_last_empty_row(sheet):
    for row in range(4, sheet.max_row + 1):
        if sheet[f"B{row}"].value is None:
            return row
    return sheet.max_row + 1

# ================================
# 计算 SUM
# ================================
def calculate_sum(sheet, formula):
    if isinstance(formula, (int, float)):
        return float(formula)
    if not isinstance(formula, str):
        return 0.0

    match = re.match(r"^=SUM\((.+)\)$", formula)
    if match:
        cell_range = match.group(1)
        start_cell, end_cell = cell_range.split(":")
        start_row = int(re.findall(r"\d+", start_cell)[0])
        start_col = openpyxl.utils.cell.column_index_from_string(re.findall(r"[A-Z]+", start_cell)[0])
        end_row = int(re.findall(r"\d+", end_cell)[0])
        total = 0.0
        for row in range(start_row, end_row + 1):
            value = sheet.cell(row=row, column=start_col).value
            if isinstance(value, (int, float)):
                total += float(value)
        return total
    return 0.0

# ================================
# 汇总信息（行尾合计行）
# ================================
def prepare_summary_text(sheet, last_empty_row):
    stock_total       = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=10).value)  # J列
    monthly_plan      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=16).value)  # P列
    plan_gap_output   = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=17).value)  # Q列
    monthly_sent      = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=18).value)  # R列
    monthly_received  = calculate_sum(sheet, sheet.cell(row=last_empty_row, column=19).value)  # S列
    monthly_remaining = (monthly_plan - monthly_sent) if (monthly_plan and monthly_sent) else 0.0

    print(f"📊 库存总量: {stock_total}, 月计划: {monthly_plan}, 缺口排产: {plan_gap_output}, 出库: {monthly_sent}, 入库: {monthly_received}")
    return stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining

# ================================
# 微信消息（纯文本）
# ================================
def _to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            return float(s)
        except Exception:
            return None
    return None

def _fmt_num(v):
    f = _to_float(v)
    return f"{f:,.1f}" if f is not None else (str(v) if v is not None else "")

def construct_wechat_message(sheet, colored_rows, date, date2,
                             stock_total, monthly_plan, plan_gap_output,
                             monthly_sent, monthly_received, monthly_remaining,
                             max_list=20):
    # 数值转 float（兼容字符串里带逗号）
    def _to_float(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.replace(",", "").strip()
            try:
                return float(s)
            except Exception:
                return None
        return None

    # 数字格式：千分位，最多1位小数；如果正好是 .0 就去掉；右对齐到给定宽度
    def _fmt_num_fixed(v, width=8):
        f = _to_float(v)
        if f is None:
            s = ""
        else:
            s = f"{f:,.1f}"
            if s.endswith(".0"):
                s = s[:-2]
        return s.rjust(width)

    # 编号：左对齐，宽度更紧
    def _fmt_code(c, width=5):
        if c is None:
            s = ""
        else:
            s = str(c).strip()
            if s.isdigit():
                s = s.zfill(5)
        return s.ljust(width)

    # 明细数据（低于 50% 外应存）
    items = []
    for r in colored_rows:
        code     = sheet.cell(row=r, column=3).value   # C 编号
        stock    = sheet.cell(row=r, column=10).value  # J 外仓库存
        expected = sheet.cell(row=r, column=11).value  # K 外应存
        home     = sheet.cell(row=r, column=13).value  # M 家库存
        items.append((code, stock, expected, home))

    lines = []
    if date:
        lines.append(f"【{date} 重庆俊都仓储数据】")
    if date2:
        lines.append(f"【{date2} 家里库存数据】")
    lines.append(f"低于 50% 外应存：{len(items)} 款")

    # 明细表（列更紧凑）
    lines.append(f"— 明细（最多展示 {max_list} 款）—")
    header = f"{'编号'.ljust(5)} {'外库'.rjust(8)} {'应存'.rjust(8)} {'家库'.rjust(8)}"
    sep    = "-" * len(header)
    lines.append(header)
    lines.append(sep)

    for code, stock, expected, home in items[:max_list]:
        line = (
            f"{_fmt_code(code, 5)} "
            f"{_fmt_num_fixed(stock, 8)} "
            f"{_fmt_num_fixed(expected, 8)} "
            f"{_fmt_num_fixed(home, 8)}"
        )
        lines.append(line)

    if len(items) > max_list:
        lines.append(f"……其余 {len(items) - max_list} 款略")

    # 汇总（放在最后；同样使用收紧后的数字格式）
    lines.append("— 汇总 —")
    lines.append(f"外仓库存总量：{_fmt_num_fixed(stock_total).strip()}")
    lines.append(f"月计划：{_fmt_num_fixed(monthly_plan).strip()}")
    lines.append(f"月计划缺口排产：{_fmt_num_fixed(plan_gap_output).strip()}")
    lines.append(f"外仓出库总量：{_fmt_num_fixed(monthly_sent).strip()}")
    lines.append(f"外仓入库总量：{_fmt_num_fixed(monthly_received).strip()}")
    lines.append(f"月预估还有要发货：{_fmt_num_fixed(monthly_remaining).strip()}")

    return "\n".join(lines)

# ================================
# 保存 TXT
# ================================
def save_wechat_message_to_file(message, output_dir):
    path = os.path.join(output_dir, "wechat_msg.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(message)
    print(f"📄 已保存微信消息文本：{path}")

# ================================
# 主程序
# ================================
def main():
    inventory_folder = get_inventory_folder()
    inventory_file = find_excel_file(inventory_folder)
    sheet = load_worksheet(inventory_file)

    colored_rows = find_colored_rows(sheet)
    date, date2 = get_dates(sheet)
    last_empty_row = find_last_empty_row(sheet)
    print(f"⚡ 发现 B 列第一个空单元格所在行: {last_empty_row}")

    stock_total, monthly_plan, plan_gap_output, monthly_sent, monthly_received, monthly_remaining = prepare_summary_text(sheet, last_empty_row)

    # 生成并保存“微信消息版”
    wechat_msg = construct_wechat_message(
        sheet, colored_rows, date, date2,
        stock_total, monthly_plan, plan_gap_output,
        monthly_sent, monthly_received, monthly_remaining,
        max_list=MAX_LIST
    )
    save_wechat_message_to_file(wechat_msg, inventory_folder)
    print("\n====== 微信消息预览 ======\n" + wechat_msg)

    print("✅ 红色或紫色单元格数量：", len(colored_rows))
    print("📌 行号列表：", colored_rows)

if __name__ == "__main__":
    main()
