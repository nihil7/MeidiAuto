import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series

# ======================== 配置区域 ========================
folder_path = r'C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\32重庆-美的\美的发货\月度汇总'
start_month = 2502
end_month = 2512
output_filename = f'月汇总表{start_month}-{end_month}.xlsx'
SHEET_NAME_IN = '入库汇总'
SHEET_NAME_OUT = '出库汇总'
SRC_SHEET = '库存表'
COL_CODE = '编号'
COL_IN = '外仓入库总量'
COL_OUT = '外仓出库总量'
# ==========================================================

def extract_month(filename):
    m = re.search(r'(\d{4})月底', filename)
    return int(m.group(1)) if m else None

def read_monthly_data(file_path, month_str, target_col):
    wb = load_workbook(file_path, data_only=True)
    if SRC_SHEET not in wb.sheetnames:
        return None
    ws = wb[SRC_SHEET]
    headers = [cell.value for cell in ws[4]]
    if COL_CODE not in headers or target_col not in headers:
        return None

    idx_code = headers.index(COL_CODE) + 1
    idx_val = headers.index(target_col) + 1

    rec = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = row[idx_code - 1]
        val = row[idx_val - 1]
        if code is not None:
            rec.append({COL_CODE: code, f'{month_str}{target_col}': val})
    return pd.DataFrame(rec) if rec else None

def merge_months(dfs):
    if not dfs:
        return None
    out = dfs[0]
    for d in dfs[1:]:
        out = pd.merge(out, d, on=COL_CODE, how='outer')
    out = out.fillna(0)

    # 增加合计列并按合计降序排序
    month_cols = [c for c in out.columns if c != COL_CODE]
    out["合计"] = out[month_cols].sum(axis=1)
    out = out.sort_values(by="合计", ascending=False)
    return out

def write_two_sheets_and_save(path, df_in, df_out):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        if not df_in.empty:
            df_in.to_excel(writer, index=False, sheet_name=SHEET_NAME_IN)
        if not df_out.empty:
            df_out.to_excel(writer, index=False, sheet_name=SHEET_NAME_OUT)

    # 添加自动筛选
    wb = load_workbook(path)
    for sheet_name in [SHEET_NAME_IN, SHEET_NAME_OUT]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row, max_col = ws.max_row, ws.max_column
            ws.auto_filter.ref = f"A1:{chr(64+max_col)}{max_row}"
    wb.save(path)

def create_excel_chart(excel_path, sheet_name, title):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    max_col, max_row = ws.max_column, ws.max_row

    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = '数量'
    chart.x_axis.title = '月份'
    chart.x_axis.number_format = 'General'
    chart.x_axis.majorTickMark = "out"

    # 去掉“合计”列，不然图表重复统计
    cats = Reference(ws, min_col=2, max_col=max_col-1, min_row=1)
    chart.set_categories(cats)

    for r in range(2, max_row + 1):
        series = Series(
            Reference(ws, min_col=2, max_col=max_col-1, min_row=r, max_row=r),
            title=str(ws.cell(row=r, column=1).value)
        )
        chart.series.append(series)

    ws.add_chart(chart, f"B{max_row + 3}")
    wb.save(excel_path)

def main():
    in_dfs, out_dfs, seen = [], [], set()

    for file in os.listdir(folder_path):
        if not file.endswith('.xlsx'):
            continue
        month = extract_month(file)
        if not month or not (start_month <= month <= end_month) or month in seen:
            continue
        seen.add(month)
        fp, mstr = os.path.join(folder_path, file), str(month)
        d_in = read_monthly_data(fp, mstr, COL_IN)
        if d_in is not None: in_dfs.append(d_in)
        d_out = read_monthly_data(fp, mstr, COL_OUT)
        if d_out is not None: out_dfs.append(d_out)

    if not in_dfs and not out_dfs:
        print("⚠️ 没有可用数据")
        return

    df_in = merge_months(in_dfs) if in_dfs else pd.DataFrame(columns=[COL_CODE])
    df_out = merge_months(out_dfs) if out_dfs else pd.DataFrame(columns=[COL_CODE])
    output_path = os.path.join(folder_path, output_filename)

    write_two_sheets_and_save(output_path, df_in, df_out)

    if not df_in.empty:
        create_excel_chart(output_path, SHEET_NAME_IN, "入库趋势")
    if not df_out.empty:
        create_excel_chart(output_path, SHEET_NAME_OUT, "出库趋势")

if __name__ == "__main__":
    main()
